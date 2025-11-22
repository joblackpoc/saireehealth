"""
Health App views with graphs, comparisons, and health summaries
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import io
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
import os
from django.contrib.auth.models import User
from .models import HealthRecord
from .forms import HealthRecordForm, DateRangeFilterForm, HealthIndicatorReportForm, HealthStatusReportForm, AdminHealthReportForm
from accounts.models import UserActivity


def register_thai_fonts():
    """Register Thai fonts for PDF rendering"""
    try:
        # Try to register Thai fonts available on the system
        font_paths = [
            r'C:\Windows\Fonts\tahoma.ttf',  # Tahoma (supports Thai)
            r'C:\Windows\Fonts\arial.ttf',   # Arial (limited Thai)
            r'C:\Windows\Fonts\THSarabunNew.ttf',  # Thai font if available
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                font_name = os.path.basename(font_path).split('.')[0]
                if font_name.lower() == 'tahoma':
                    pdfmetrics.registerFont(TTFont('Tahoma', font_path))
                    return 'Tahoma'
                elif font_name.lower() == 'arial':
                    pdfmetrics.registerFont(TTFont('Arial', font_path))
                    return 'Arial'
                elif font_name.lower() == 'thsarabunnew':
                    pdfmetrics.registerFont(TTFont('THSarabun', font_path))
                    return 'THSarabun'
    except Exception as e:
        print(f"Font registration error: {e}")
    
    # Fallback to Helvetica
    return 'Helvetica'

def safe_thai_text(text):
    """Safely handle Thai text for PDF rendering"""
    if not text:
        return ""
    
    try:
        # Convert to string first
        text_str = str(text)
        
        # Handle Unicode properly - remove surrogates if any
        clean_text = text_str.encode('utf-8', errors='ignore').decode('utf-8')
        
        # Return the cleaned text
        return clean_text
    except Exception:
        # Fallback to ASCII-safe version
        try:
            return str(text).encode('ascii', errors='ignore').decode('ascii')
        except Exception:
            return "Text encoding error"


def log_health_activity(user, action, description='', request=None):
    """Log health-related activity"""
    activity = UserActivity.objects.create(
        user=user,
        action=action,
        description=description
    )
    if request:
        activity.ip_address = request.META.get('REMOTE_ADDR', '')
        activity.user_agent = request.META.get('HTTP_USER_AGENT', '')
        activity.save()


@login_required
def dashboard_view(request):
    """Main dashboard with health overview"""
    user = request.user
    profile = user.profile
    
    # Check if MFA is required but not setup
    if profile.requires_mfa() and not profile.mfa_enabled:
        messages.error(request, '⚠️ คุณต้องตั้งค่า MFA ก่อนใช้งานระบบ')
        return redirect('accounts:mfa_setup')
    
    # Check if profile is complete
    if not profile.firstname or not profile.lastname or not profile.age:
        messages.warning(request, 'กรุณาเพิ่มข้อมูลโปรไฟล์ให้ครบถ้วนก่อนใช้งาน')
        return redirect('accounts:profile')
    
    # Get latest record
    latest_record = HealthRecord.objects.filter(user=user).first()
    
    if not latest_record:
        messages.info(request, 'กรุณาเพิ่มข้อมูลสุขภาพของคุณ')
        return redirect('health_app:add_metric')
    
    # Get status for all metrics
    bmi_status = latest_record.get_bmi_status()
    fat_status = latest_record.get_fat_percent_status()
    visceral_status = latest_record.get_visceral_fat_status()
    muscle_status = latest_record.get_muscle_percent_status()
    blood_pressure_status = latest_record.get_blood_pressure_status()
    waist_status = latest_record.get_waist_status()
    cholesterol_status = latest_record.get_cholesterol_status()
    ldl_status = latest_record.get_ldl_status()
    hdl_status = latest_record.get_hdl_status()
    fbs_status = latest_record.get_fbs_status()
    triglycerides_status = latest_record.get_triglycerides_status()
    
    # Generate health overview recommendations
    health_overview = generate_health_overview(latest_record, profile, bmi_status, fat_status, visceral_status, muscle_status)
    
    context = {
        'profile': profile,
        'latest_record': latest_record,
        'bmi_status': bmi_status,
        'fat_status': fat_status,
        'visceral_status': visceral_status,
        'muscle_status': muscle_status,
        'blood_pressure_status': blood_pressure_status,
        'waist_status': waist_status,
        'cholesterol_status': cholesterol_status,
        'ldl_status': ldl_status,
        'hdl_status': hdl_status,
        'fbs_status': fbs_status,
        'triglycerides_status': triglycerides_status,
        'health_overview': health_overview,
    }
    
    return render(request, 'health_app/dashboard_n.html', context)


def generate_health_overview(record, profile, bmi_status, fat_status, visceral_status, muscle_status):

    def create_item(text, color='info', severity='normal'):
        """Create overview item with color"""
        return {
            'text': text,
            'color': color,      # Bootstrap alert class
            'severity': severity
        }
    
    """Generate comprehensive health overview with recommendations"""
    overview = []
    
    # Blood Pressure Recommendations (with detailed conditions and advice)
    systolic = record.blood_pressure_systolic
    diastolic = record.blood_pressure_diastolic
    
    if systolic < 120 and diastolic < 80:
        overview.append(create_item('คุณมีความดันโลหิตในระดับเหมาะสม', 'success'))
        overview.append(create_item('คำแนะนำ: ควบคุมอาหาร, มีกิจกรรมทางกาย และวัดความดันสม่ำเสมอ', 'success'))
    elif 120 <= systolic <= 129 and 80 <= diastolic <= 84:
        overview.append(create_item('คุณมีความดันโลหิตในระดับปกติ', 'success'))
        overview.append(create_item('คำแนะนำ: ควบคุมอาหาร, มีกิจกรรมทางกาย และวัดความดันสม่ำเสมอ', 'success'))
    elif (130 <= systolic <= 139) or (85 <= diastolic <= 89):
        overview.append(create_item('คุณมีความดันโลหิตสูงกว่าปกติ', 'warning'))
        overview.append(create_item('คำแนะนำ: ลดน้ำหนักหากมีน้ำหนักเกิน, หลีกเลี่ยงความเครียด บุหรี่, มีกิจกรรมทางกายอย่างสม่ำเสมอ, ลดการกินเค็ม', 'warning'))
    elif (140 <= systolic <= 159) or (90 <= diastolic <= 99):
        overview.append(create_item('คุณอาจเป็นโรคความดันโลหิตสูงระดับที่ 1', 'warning'))
        overview.append(create_item('คำแนะนำ: ควรรีบปรึกษาแพทย์เพื่อรับการวินิจฉัยและรับการรักษาที่เหมาะสม รวมถึงเข้าสู่กระบวนการปรับเปลี่ยนพฤติกรรม', 'warning'))
    elif (160 <= systolic <= 179) or (100 <= diastolic <= 109):
        overview.append(create_item('คุณอาจเป็นโรคความดันโลหิตสูงระดับที่ 2', 'warning'))
        overview.append(create_item('คำแนะนำ: ควรรีบปรึกษาแพทย์เพื่อรับการวินิจฉัยและรับการรักษาที่เหมาะสม รวมถึงเข้าสู่กระบวนการปรับเปลี่ยนพฤติกรรม', 'warning'))
    elif systolic >= 180 or diastolic >= 110:
        overview.append(create_item('คุณอาจเป็นโรคความดันโลหิตสูงระดับที่ 3', 'danger'))
        overview.append(create_item('คำแนะนำ: ควรรีบพบแพทย์ทันที', 'danger'))
    elif systolic >= 140 and diastolic < 90:
        overview.append(create_item('คุณอาจเป็นโรคความดันโลหิตสูงเฉพาะตัวบน', 'danger'))
        overview.append(create_item('คำแนะนำ: ควรปรึกษาบุคลากรทางการแพทย์', 'danger'))
    elif systolic < 140 and diastolic >= 90:
        overview.append(create_item('คุณอาจเป็นโรคความดันโลหิตสูงเฉพาะตัวล่าง', 'danger'))
        overview.append(create_item('คำแนะนำ: ควรปรึกษาบุคลากรทางการแพทย์', 'danger'))
    
    # Waist Circumference (calculated from height)
    waist = float(record.waist)
    height = float(record.height)
    waist_threshold = height / 2
    
    if abs(waist - waist_threshold) < 1:  # Approximately equal
        overview.append(create_item('เส้นรอบเอวของคุณอยู่ในเกณฑ์ปกติ', 'success'))
    elif waist > waist_threshold:
        overview.append(create_item('เส้นรอบเอวของคุณเกินเกณฑ์', 'danger'))
    else:  # waist < waist_threshold
        overview.append(create_item('รอบเอวของคุณอยู่ในเกณฑ์ดีมาก', 'success'))
    
    # BMI Recommendations (following exact conditions)
    bmi = float(record.bmi)
    
    if bmi < 18.5:
        overview.append(create_item('น้อยกว่าปกติ/ผอม - ภาวะเสี่ยงต่อโรค มากกว่าคนปกติ', 'warning'))
    elif 18.5 <= bmi <= 22.9:
        overview.append(create_item('ปกติ/สุขภาพดี - ภาวะเสี่ยงต่อโรค เท่ากับคนปกติ', 'success'))
    elif 23 <= bmi <= 24.9:
        overview.append(create_item('ท้วม มีภาวะน้ำหนักเกินหรือโรคอ้วนระดับ 1 แนะนำให้ดูมวลกล้ามเนื้อประกอบด้วย - ภาวะเสี่ยงต่อโรค อันตรายระดับ 1 ', 'warning'))
    elif 25 <= bmi <= 30:
        overview.append(create_item('อ้วน มีภาวะน้ำหนักเกินหรือโรคอ้วนระดับ 2 - ภาวะเสี่ยงต่อโรค อันตรายระดับ 2 ', 'warning'))
    else:  # bmi > 30
        overview.append(create_item('อ้วนมาก มีภาวะน้ำหนักเกินมากอย่างมากหรือโรคอ้วนระดับ 3 - ภาวะเสี่ยงต่อโรค อันตรายระดับ 3', 'danger'))
    
    # Fat Percent Recommendations (gender-based)
    fat = float(record.fat_percent)
    gender = profile.gender
    
    if gender == 'female':
        if 5 <= fat <= 19.9:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: ต่ำ', 'success'))
        elif 20 <= fat <= 29.9:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: ปกติ', 'success'))
        elif 30 <= fat <= 34.5:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: เริ่มอ้วน ', 'warning'))
        elif 35 <= fat <= 50:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: อ้วน - คุณมีไขมันในร่างกายมากเกินไป ', 'danger'))
    else:  # male
        if 5 <= fat <= 9.9:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: ต่ำ', 'success'))
        elif 10 <= fat <= 19.9:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: ปกติ', 'success'))
        elif 20 <= fat <= 24.9:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: เริ่มอ้วน ', 'warning'))
        elif 25 <= fat <= 50:
            overview.append(create_item('เปอร์เซ็นต์ไขมัน: อ้วน - คุณมีไขมันในร่างกายมากเกินไป ', 'danger'))
    # Visceral Fat Recommendations
    vf = float(record.visceral_fat)
    
    if 1 <= vf <= 9:
        overview.append(create_item('ไขมันในช่องท้อง: ปกติ', 'success'))
    elif 10 <= vf <= 14:
        overview.append(create_item('ไขมันในช่องท้อง: สูง - คุณมีภาวะเสี่ยงจากการมีไขมันช่องท้องสูง', 'warning'))
    elif 15 <= vf <= 30:
        overview.append(create_item('ไขมันในช่องท้อง: สูงมาก - คุณมีภาวะเสี่ยงอันตรายจากการมีไขมันช่องท้องมากผิดปกติ', 'danger'))
    
    # Muscle Percent Recommendations (age and gender-based)
    muscle = float(record.muscle_percent)
    age = profile.age
    
    if gender == 'female':
        if 18 <= age <= 39:
            if muscle < 24.3:
                overview.append(create_item('มวลกล้ามเนื้อ: ต่ำ - คุณมีมวลกล้ามเนื้อน้อยเกินไป สร้างกล้ามเนื้อโดยเน้น เพิ่มโปรตีนคุณภาพ นอนหลับคุณภาพ ออกกำลังกายแบบบอดี้เวท หรือเวทเทรนนิ่ง', 'warning'))
            elif 24.3 <= muscle <= 30.3:
                overview.append(create_item('มวลกล้ามเนื้อ: ปกติ', 'success'))
            elif 30.4 <= muscle <= 35.3:
                overview.append(create_item('มวลกล้ามเนื้อ: สูง ร่างกายแข็งแรง', 'success'))
            elif muscle >= 35.4:
                overview.append(create_item('มวลกล้ามเนื้อ: สูงมาก - คุณมีมวลกล้ามเนื้อในระดับนักกีฬา ร่างกายแข็งแรง พยายามรักษาพฤติกรรมสุขภาพต่อไป', 'success','success'))
        elif 40 <= age <= 59:
            if muscle < 24.1:
                overview.append(create_item('มวลกล้ามเนื้อ: ต่ำ - คุณมีมวลกล้ามเนื้อน้อยเกินไป สร้างกล้ามเนื้อโดยเน้น เพิ่มโปรตีนคุณภาพ นอนหลับคุณภาพ ออกกำลังกายแบบบอดี้เวท หรือเวทเทรนนิ่ง', 'warning'))
            elif 24.1 <= muscle <= 30.1:
                overview.append(create_item('มวลกล้ามเนื้อ: ปกติ', 'success'))
            elif 30.2 <= muscle <= 35.1:
                overview.append(create_item('มวลกล้ามเนื้อ: สูง ร่างกายแข็งแรง', 'success'))
            elif muscle >= 35.2:
                overview.append(create_item('มวลกล้ามเนื้อ: สูงมาก - คุณมีมวลกล้ามเนื้อในระดับนักกีฬา ร่างกายแข็งแรง พยายามรักษาพฤติกรรมสุขภาพต่อไป', 'success','success'))
        elif 60 <= age <= 80:
            if muscle < 23.9:
                overview.append(create_item('มวลกล้ามเนื้อ: ต่ำ - คุณมีมวลกล้ามเนื้อน้อยเกินไป สร้างกล้ามเนื้อโดยเน้น เพิ่มโปรตีนคุณภาพ นอนหลับคุณภาพ ออกกำลังกายแบบบอดี้เวท หรือเวทเทรนนิ่ง', 'warning'))
            elif 23.9 <= muscle <= 29.9:
                overview.append(create_item('มวลกล้ามเนื้อ: ปกติ', 'success'))
            elif 30 <= muscle <= 34.9:
                overview.append(create_item('มวลกล้ามเนื้อ: สูง ร่างกายแข็งแรง', 'success'))
            elif muscle >= 35:
                overview.append(create_item('มวลกล้ามเนื้อ: สูงมาก - คุณมีมวลกล้ามเนื้อในระดับนักกีฬา ร่างกายแข็งแรง พยายามรักษาพฤติกรรมสุขภาพต่อไป', 'success','success'))
    else:  # male
        if 18 <= age <= 39:
            if muscle < 33.3:
                overview.append(create_item('มวลกล้ามเนื้อ: ต่ำ - คุณมีมวลกล้ามเนื้อน้อยเกินไป สร้างกล้ามเนื้อโดยเน้น เพิ่มโปรตีนคุณภาพ นอนหลับคุณภาพ ออกกำลังกายแบบบอดี้เวท หรือเวทเทรนนิ่ง', 'warning'))
            elif 33.3 <= muscle <= 39.3:
                overview.append(create_item('มวลกล้ามเนื้อ: ปกติ', 'success'))
            elif 39.4 <= muscle <= 44:
                overview.append(create_item('มวลกล้ามเนื้อ: สูง ร่างกายแข็งแรง', 'success'))
            elif muscle >= 44.1:
                overview.append(create_item('มวลกล้ามเนื้อ: สูงมาก - คุณมีมวลกล้ามเนื้อในระดับนักกีฬา ร่างกายแข็งแรง พยายามรักษาพฤติกรรมสุขภาพต่อไป', 'success','success'))
        elif 40 <= age <= 59:
            if muscle < 33.1:
                overview.append(create_item('มวลกล้ามเนื้อ: ต่ำ - คุณมีมวลกล้ามเนื้อน้อยเกินไป สร้างกล้ามเนื้อโดยเน้น เพิ่มโปรตีนคุณภาพ นอนหลับคุณภาพ ออกกำลังกายแบบบอดี้เวท หรือเวทเทรนนิ่ง', 'warning'))
            elif 33.1 <= muscle <= 39.1:
                overview.append(create_item('มวลกล้ามเนื้อ: ปกติ', 'success'))
            elif 39.2 <= muscle <= 43.8:
                overview.append(create_item('มวลกล้ามเนื้อ: สูง ร่างกายแข็งแรง', 'success'))
            elif muscle >= 43.9:
                overview.append(create_item('มวลกล้ามเนื้อ: สูงมาก - คุณมีมวลกล้ามเนื้อในระดับนักกีฬา ร่างกายแข็งแรง พยายามรักษาพฤติกรรมสุขภาพต่อไป', 'success','success'))
        elif 60 <= age <= 80:
            if muscle < 32.9:
                overview.append(create_item('มวลกล้ามเนื้อ: ต่ำ - คุณมีมวลกล้ามเนื้อน้อยเกินไป สร้างกล้ามเนื้อโดยเน้น เพิ่มโปรตีนคุณภาพ นอนหลับคุณภาพ ออกกำลังกายแบบบอดี้เวท หรือเวทเทรนนิ่ง', 'warning'))
            elif 32.9 <= muscle <= 38.9:
                overview.append(create_item('มวลกล้ามเนื้อ: ปกติ', 'success'))
            elif 39 <= muscle <= 43.6:
                overview.append(create_item('มวลกล้ามเนื้อ: สูง ร่างกายแข็งแรง', 'success'))
            elif muscle >= 43.7:
                overview.append(create_item('มวลกล้ามเนื้อ: สูงมาก - คุณมีมวลกล้ามเนื้อในระดับนักกีฬา ร่างกายแข็งแรง พยายามรักษาพฤติกรรมสุขภาพต่อไป', 'success','success'))
    
    # Cholesterol (if available)
    if record.cholesterol:
        chol = float(record.cholesterol)
        if chol < 200:
            overview.append(create_item('ปริมาณไขมันคอเลสเตอรอลอยู่ในระดับปกติ', 'success'))
        else:  # chol > 200
            overview.append(create_item('ปริมาณไขมันคอเลสเตอรอลอยู่ในระดับมากผิดปกติ', 'warning'))
    
    # LDL (if available)
    if record.ldl:
        ldl = float(record.ldl)
        if ldl < 130:
            overview.append(create_item('ไขมันไม่ดีอยู่ในระดับปกติ', 'success'))
        else:  # ldl > 130
            overview.append(create_item('ปริมาณไขมันไม่ดีอยู่ในระดับมากผิดปกติ', 'warning'))
    
    # HDL (if available)
    if record.hdl:
        hdl = float(record.hdl)
        if gender == 'male':
            if hdl > 40:
                overview.append(create_item('ปริมาณไขมันดีอยู่ในระดับปกติ', 'success'))
            else:  # hdl < 40
                overview.append(create_item('ปริมาณไขมันดีอยู่ในระดับต่ำกว่าปกติ ', 'warning'))
        else:  # female
            if hdl > 50:
                overview.append(create_item('ปริมาณไขมันดีอยู่ในระดับปกติ', 'success'))
            else:  # hdl < 50
                overview.append(create_item('ปริมาณไขมันดีอยู่ในระดับต่ำกว่าปกติ ', 'warning'))
    
    # FBS (if available)
    if record.fbs:
        fbs = float(record.fbs)
        if fbs < 100:
            overview.append(create_item('ระดับน้ำตาลอยู่ในเกณฑ์ปกติ', 'success'))
        elif 100 <= fbs <= 125:
            overview.append(create_item('ระดับน้ำตาลในเลือดเสี่ยงเป็นโรคเบาหวานหรือมีภาวะก่อนเบาหวาน', 'warning'))
        else:  # fbs >= 126
            overview.append(create_item('ระดับน้ำตาลในเลือดอยู่ในเกณฑ์โรคเบาหวาน', 'danger'))
    # Triglycerides (if available)
    if record.triglycerides:
        tg = float(record.triglycerides)
        if tg < 150:
            overview.append(create_item('ระดับไตรกลีเซอไรด์อยู่ในเกณฑ์ปกติ', 'success'))
        elif 150 <= tg <= 199:
            overview.append(create_item('ระดับไตรกลีเซอไรด์อยู่ในเกณฑ์สูงเล็กน้อย ลดอาหารที่มีน้ำตาลและไขมันทรานส์ กินอาหารคาร์บคุณภาพ งด/ลดอาหารแปรรูปขั้นสูง งดเครื่องดื่มแอลกอฮอล์ และทำ IF กินและอดอาหารเป็นช่วงๆ', 'warning'))
        elif 200 <= tg <= 499:
            overview.append(create_item('ระดับไตรกลีเซอไรด์อยู่ในเกณฑ์สูง ลดอาหารที่มีน้ำตาลและไขมันทรานส์ กินอาหารคาร์บคุณภาพ งด/ลดอาหารแปรรูปขั้นสูง งดเครื่องดื่มแอลกอฮอล์ และทำ IF กินและอดอาหารเป็นช่วงๆ ', 'warning'))
        else:  # tg >= 500
            overview.append(create_item('ระดับไตรกลีเซอไรด์อยู่ในเกณฑ์สูงมาก ซึ่งอาจเพิ่มความเสี่ยงต่อการเกิดตับอ่อนอักเสบ ควรรีบปรึกษาบุคลากรทางการแพทย์ทันที', 'danger'))

    return overview



@login_required
def add_metric_view(request):
    """Add new health metric"""
    profile = request.user.profile
    
    # Check if profile is complete
    if not profile.firstname or not profile.lastname or not profile.age:
        messages.warning(request, 'กรุณาเพิ่มข้อมูลโปรไฟล์ให้ครบถ้วนก่อนเพิ่มข้อมูลสุขภาพ')
        return redirect('accounts:profile')
    
    if request.method == 'POST':
        form = HealthRecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            record.user = request.user
            record.save()
            log_health_activity(request.user, 'health_record_add', f'Added health record for {record.recorded_at}', request)
            messages.success(request, 'เพิ่มข้อมูลสุขภาพสำเร็จ!')
            return redirect('health_app:dashboard')
    else:
        form = HealthRecordForm()
    
    context = {
        'form': form,
        'profile': profile,
    }
    
    return render(request, 'health_app/add_metric.html', context)


@login_required
@login_required
@login_required
def update_record_view(request, record_id):
    """Update existing health record"""
    record = get_object_or_404(HealthRecord, id=record_id, user=request.user)
    
    if request.method == 'POST':
        form = HealthRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            log_health_activity(request.user, 'health_record_update', f'Updated health record for {record.recorded_at}', request)
            messages.success(request, 'อัปเดตข้อมูลสำเร็จ!')
            return redirect('health_app:history')
    else:
        form = HealthRecordForm(instance=record)
    
    return render(request, 'health_app/update_record.html', {'form': form, 'record': record})


@login_required
def history_view(request):
    """History view with graphs and comparison tables"""
    user = request.user
    
    # Get all records
    records = HealthRecord.objects.filter(user=user).order_by('recorded_at')
    
    if not records.exists():
        messages.info(request, 'ไม่มีข้อมูลประวัติ')
        return redirect('health_app:add_metric')
    
    # Get filter parameters
    metric_filter = request.GET.get('metric', 'bmi')  # Default to BMI
    date_start_str = request.GET.get('date_start', '')
    date_end_str = request.GET.get('date_end', '')
    
    # Apply date filters
    filtered_records = records
    if date_start_str:
        date_start = datetime.strptime(date_start_str, '%Y-%m-%d').date()
        filtered_records = filtered_records.filter(recorded_at__date__gte=date_start)
    if date_end_str:
        date_end = datetime.strptime(date_end_str, '%Y-%m-%d').date()
        filtered_records = filtered_records.filter(recorded_at__date__lte=date_end)
    
    # Get first and last dates for defaults
    first_record = records.first()
    last_record = records.last()
    
    # Generate graph
    graph_base64 = generate_metric_graph(filtered_records, metric_filter, user.profile)
    
    # Generate comparison table
    comparison_data = None
    summary_info = None
    if filtered_records.count() >= 2:
        first_filtered = filtered_records.first()
        last_filtered = filtered_records.last()
        comparison_data = generate_comparison_table(first_filtered, last_filtered, user.profile)
        summary_info = generate_summary_info(first_filtered, last_filtered, user.profile)
    
    context = {
        'records': records,
        'graph_base64': graph_base64,
        'metric_filter': metric_filter,
        'date_start': date_start_str or (first_record.recorded_at.strftime('%Y-%m-%d') if first_record else ''),
        'date_end': date_end_str or (last_record.recorded_at.strftime('%Y-%m-%d') if last_record else ''),
        'first_record': first_record,
        'last_record': last_record,
        'comparison_data': comparison_data,
        'summary_info': summary_info,
    }
    
    return render(request, 'health_app/history.html', context)


def generate_metric_graph(records, metric_name, profile):
    """Generate line graph for specific metric"""
    if not records.exists():
        return None
    
    plt.figure(figsize=(12, 6))
    plt.rcParams['font.family'] = 'DejaVu Sans'
    
    dates = [r.recorded_at for r in records]
    
    # Metric configuration
    metric_config = {
        'bmi': {
            'values': [float(r.bmi) for r in records],
            'title': 'BMI',
            'ylabel': 'BMI',
            'normal_range': (18.5, 22.9),
        },
        'fat_percent': {
            'values': [float(r.fat_percent) for r in records],
            'title': 'Fat Percent',
            'ylabel': 'Fat Percent (%)',
            'normal_range': get_fat_normal_range(profile),
        },
        'visceral_fat': {
            'values': [float(r.visceral_fat) for r in records],
            'title': 'Visceral Fat',
            'ylabel': 'Visceral Fat',
            'normal_range': (1, 9),
        },
        'muscle_percent': {
            'values': [float(r.muscle_percent) for r in records],
            'title': 'Muscle Percent',
            'ylabel': 'Muscle Percent (%)',
            'normal_range': get_muscle_normal_range(profile),
        },
        'blood_pressure_systolic': {
            'values': [float(r.blood_pressure_systolic) for r in records],
            'title': 'Blood Pressure Systolic',
            'ylabel': 'Systolic (mmHg)',
            'normal_range': (90, 120),
        },
        'blood_pressure_diastolic': {
            'values': [float(r.blood_pressure_diastolic) for r in records],
            'title': 'Blood Pressure Diastolic',
            'ylabel': 'Diastolic (mmHg)',
            'normal_range': (60, 80),
        },
        'waist': {
            'values': [float(r.waist) for r in records],
            'title': 'Waist Circumference',
            'ylabel': 'Waist (cm)',
            'normal_range': (60, 90 if profile.gender == 'male' else 80),
        },
        'cholesterol': {
            'values': [float(r.cholesterol) for r in records if r.cholesterol],
            'title': 'Cholesterol',
            'ylabel': 'Cholesterol (mg/dL)',
            'normal_range': (0, 200),
        },
        'ldl': {
            'values': [float(r.ldl) for r in records if r.ldl],
            'title': 'LDL',
            'ylabel': 'LDL (mg/dL)',
            'normal_range': (0, 100),
        },
        'hdl': {
            'values': [float(r.hdl) for r in records if r.hdl],
            'title': 'HDL',
            'ylabel': 'HDL (mg/dL)',
            'normal_range': (40, 100),
        },
        'fbs': {
            'values': [float(r.fbs) for r in records if r.fbs],
            'title': 'FBS (Fasting Blood Sugar)',
            'ylabel': 'FBS (mg/dL)',
            'normal_range': (70, 100),
        },
        'triglycerides': {
            'values': [float(r.triglycerides) for r in records if r.triglycerides],
            'title': 'Triglycerides',
            'ylabel': 'Triglycerides (mg/dL)',
            'normal_range': (0, 150),
        },
    }
    
    config = metric_config.get(metric_name, metric_config['bmi'])
    values = config['values']
    normal_range = config['normal_range']
    
    # Plot values
    plt.plot(dates, values, marker='o', linestyle='-', linewidth=2, markersize=8, label='Your Values')
    
    # Plot normal range as green horizontal lines
    if normal_range:
        plt.axhline(y=normal_range[0], color='green', linestyle='--', linewidth=2, alpha=0.7, label=f'Normal Range ({normal_range[0]}-{normal_range[1]})')
        plt.axhline(y=normal_range[1], color='green', linestyle='--', linewidth=2, alpha=0.7)
        plt.fill_between(dates, normal_range[0], normal_range[1], color='green', alpha=0.1)
    
    plt.title(config['title'], fontsize=16, fontweight='bold')
    plt.xlabel('Date', fontsize=12)
    plt.ylabel(config['ylabel'], fontsize=12)
    plt.xticks(rotation=45)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    
    # Convert to base64
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    
    return image_base64


def get_fat_normal_range(profile):
    """Get normal fat percent range based on gender"""
    if profile.gender == 'female':
        return (20, 29.9)
    else:
        return (10, 19.9)


def get_muscle_normal_range(profile):
    """Get normal muscle percent range based on age and gender"""
    age = profile.age
    gender = profile.gender
    
    if gender == 'female':
        if 18 <= age <= 39:
            return (24.3, 30.3)
        elif 40 <= age <= 59:
            return (24.1, 30.1)
        else:  # 60-80
            return (23.9, 29.9)
    else:  # male
        if 18 <= age <= 39:
            return (33.3, 39.3)
        elif 40 <= age <= 59:
            return (33.1, 39.1)
        else:  # 60-80
            return (32.9, 38.9)
        
def get_triglycerides_status(value):
    """Get triglycerides status"""
    if value is None:
        return None, None
    if value < 150:
        return 'ปกติ', '#28a745'
    elif 150 <= value < 200:
        return 'สูงกว่าปกติ', '#ffc107'
    elif 200 <= value < 500:
        return 'สูง', '#fd7e14'
    else:
        return 'สูงมาก', '#dc3545'


def generate_comparison_table(first_record, last_record, profile):
    """Generate comparison table between two records"""
    comparison = {}
    
    # BMI comparison
    bmi_diff = float(last_record.bmi) - float(first_record.bmi)
    bmi_text = get_comparison_text(bmi_diff)
    bmi_status = last_record.get_bmi_status()
    comparison['bmi'] = {
        'start_value': float(first_record.bmi),
        'end_value': float(last_record.bmi),
        'diff': bmi_diff,
        'text': bmi_text,
        'color': bmi_status['color'],
        'normal_range': bmi_status['normal_range'],
    }
    
    # Fat percent comparison
    fat_diff = float(last_record.fat_percent) - float(first_record.fat_percent)
    fat_text = get_comparison_text(fat_diff)
    fat_status = last_record.get_fat_percent_status()
    comparison['fat_percent'] = {
        'start_value': float(first_record.fat_percent),
        'end_value': float(last_record.fat_percent),
        'diff': fat_diff,
        'text': fat_text,
        'color': fat_status['color'],
        'normal_range': fat_status['normal_range'],
    }
    
    # Visceral fat comparison
    vf_diff = float(last_record.visceral_fat) - float(first_record.visceral_fat)
    vf_text = get_comparison_text(vf_diff)
    vf_status = last_record.get_visceral_fat_status()
    comparison['visceral_fat'] = {
        'start_value': float(first_record.visceral_fat),
        'end_value': float(last_record.visceral_fat),
        'diff': vf_diff,
        'text': vf_text,
        'color': vf_status['color'],
        'normal_range': vf_status['normal_range'],
    }
    
    # Muscle percent comparison
    muscle_diff = float(last_record.muscle_percent) - float(first_record.muscle_percent)
    muscle_text = get_comparison_text(muscle_diff)
    muscle_status = last_record.get_muscle_percent_status()
    comparison['muscle_percent'] = {
        'start_value': float(first_record.muscle_percent),
        'end_value': float(last_record.muscle_percent),
        'diff': muscle_diff,
        'text': muscle_text,
        'color': muscle_status['color'],
        'normal_range': muscle_status['normal_range'],
    }
    
    # Blood pressure systolic comparison
    bp_sys_diff = last_record.blood_pressure_systolic - first_record.blood_pressure_systolic
    bp_sys_text = get_comparison_text(bp_sys_diff)
    bp_sys_status = last_record.get_blood_pressure_status()
    comparison['blood_pressure_systolic'] = {
        'start_value': first_record.blood_pressure_systolic,
        'end_value': last_record.blood_pressure_systolic,
        'diff': bp_sys_diff,
        'text': bp_sys_text,
        #'color': '#28a745' if abs(bp_sys_diff) <= 5 else '#ffc107',
        'color': bp_sys_status['color'],
        'normal_range': bp_sys_status['normal_range'],
    }
    
    # Blood pressure diastolic comparison
    bp_dia_diff = last_record.blood_pressure_diastolic - first_record.blood_pressure_diastolic
    bp_dia_text = get_comparison_text(bp_dia_diff)
    bp_dia_status = last_record.get_blood_pressure_status()
    comparison['blood_pressure_diastolic'] = {
        'start_value': first_record.blood_pressure_diastolic,
        'end_value': last_record.blood_pressure_diastolic,
        'diff': bp_dia_diff,
        'text': bp_dia_text,
        #'color': '#28a745' if abs(bp_dia_diff) <= 5 else '#ffc107',
        'color': bp_dia_status['color'],
        'normal_range': bp_dia_status['normal_range'],
    }
    
    # Waist comparison
    waist_diff = float(last_record.waist) - float(first_record.waist)
    waist_text = get_comparison_text(waist_diff)
    waist_status = last_record.get_waist_status()
    comparison['waist'] = {
        'start_value': float(first_record.waist),
        'end_value': float(last_record.waist),
        'diff': waist_diff,
        'text': waist_text,
        'color': waist_status['color'],
        'normal_range': waist_status['normal_range'],
    }
    
    # Cholesterol comparison
    if first_record.cholesterol and last_record.cholesterol:
        chol_diff = float(last_record.cholesterol) - float(first_record.cholesterol)
        chol_text = get_comparison_text(chol_diff)
        chol_status = last_record.get_cholesterol_status()
        comparison['cholesterol'] = {
            'start_value': float(first_record.cholesterol),
            'end_value': float(last_record.cholesterol),
            'diff': chol_diff,
            'text': chol_text,
            'color': chol_status['color'] if chol_status else '#6c757d',
        }
    
    # LDL comparison
    if first_record.ldl and last_record.ldl:
        ldl_diff = float(last_record.ldl) - float(first_record.ldl)
        ldl_text = get_comparison_text(ldl_diff)
        ldl_status = last_record.get_ldl_status()
        comparison['ldl'] = {
            'start_value': float(first_record.ldl),
            'end_value': float(last_record.ldl),
            'diff': ldl_diff,
            'text': ldl_text,
            'color': ldl_status['color'] if ldl_status else '#6c757d',
        }
    
    # HDL comparison
    if first_record.hdl and last_record.hdl:
        hdl_diff = float(last_record.hdl) - float(first_record.hdl)
        hdl_text = get_comparison_text(hdl_diff)
        hdl_status = last_record.get_hdl_status()
        comparison['hdl'] = {
            'start_value': float(first_record.hdl),
            'end_value': float(last_record.hdl),
            'diff': hdl_diff,
            'text': hdl_text,
            'color': hdl_status['color'] if hdl_status else '#6c757d',
        }
    
    # FBS comparison
    if first_record.fbs and last_record.fbs:
        fbs_diff = float(last_record.fbs) - float(first_record.fbs)
        fbs_text = get_comparison_text(fbs_diff)
        fbs_status = last_record.get_fbs_status()
        comparison['fbs'] = {
            'start_value': float(first_record.fbs),
            'end_value': float(last_record.fbs),
            'diff': fbs_diff,
            'text': fbs_text,
            'color': fbs_status['color'] if fbs_status else '#6c757d',
        }

    # Triglycerides comparison
    if first_record.triglycerides and last_record.triglycerides:
        tg_diff = float(last_record.triglycerides) - float(first_record.triglycerides)
        tg_text = get_comparison_text(tg_diff)
        tg_status = last_record.get_triglycerides_status()
        comparison['triglycerides'] = {
            'start_value': float(first_record.triglycerides),
            'end_value': float(last_record.triglycerides),
            'diff': tg_diff,
            'text': tg_text,
            'color': tg_status['color'] if tg_status else '#6c757d',
        }
    
    return comparison


def get_comparison_text(diff):
    """Get Thai comparison text based on difference"""
    if diff > 0:
        return f"เพิ่มขึ้น {abs(diff):.1f}"
    elif diff < 0:
        return f"ลดลง {abs(diff):.1f}"
    else:
        return "เท่าเดิม"


def generate_summary_info(first_record, last_record, profile):
    """Generate summary information with Thai text"""
    comparison = generate_comparison_table(first_record, last_record, profile)
    
    summary = {
        
        'bmi': f"BMI {comparison['bmi']['text']} ค่าปกติ {comparison['bmi']['normal_range']}",
        'waist': f"รอบเอว {comparison['waist']['text']} ค่าปกติ {comparison['waist']['normal_range']}",
        'blood_pressure_systolic': f"ความดันโลหิตตัวบน {comparison['blood_pressure_systolic']['text']} ค่าปกติ {comparison['blood_pressure_systolic']['normal_range']}",
        'blood_pressure_diastolic': f"ความดันโลหิตตัวล่าง {comparison['blood_pressure_diastolic']['text']} ค่าปกติ {comparison['blood_pressure_diastolic']['normal_range']}",
        'fat_percent': f"เปอร์เซ็นต์ไขมัน {comparison['fat_percent']['text']} ค่าปกติ {comparison['fat_percent']['normal_range']}",
        'visceral_fat': f"เปอร์เซ็นต์ไขมันในช่องท้อง {comparison['visceral_fat']['text']} ค่าปกติ {comparison['visceral_fat']['normal_range']}",
        'muscle_percent': f"เปอร์เซ็นต์มวลกล้ามเนื้อ {comparison['muscle_percent']['text']} ค่าปกติ {comparison['muscle_percent']['normal_range']}",
        'cholesterol': f"คอเลสเตอรอล {comparison['cholesterol']['text']}" if 'cholesterol' in comparison else '',
        'ldl': f"LDL {comparison['ldl']['text']}" if 'ldl' in comparison else '',
        'hdl': f"HDL {comparison['hdl']['text']}" if 'hdl' in comparison else '',
        'fbs': f"น้ำตาลในเลือดขณะอดอาหาร {comparison['fbs']['text']}" if 'fbs' in comparison else '',
        'triglycerides': f"ไตรกลีเซอไรด์ {comparison['triglycerides']['text']}" if 'triglycerides' in comparison else '',
    }
    
    return summary


@login_required
def health_report_view(request):
    """Health indicator report with filtering"""
    # Check if user has admin privileges
    if not request.user.is_staff and not request.user.profile.role in ['admin', 'superuser']:
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงรายงานสุขภาพ')
        return redirect('health_app:dashboard')
    
    form = HealthIndicatorReportForm()
    filtered_users = []
    filter_applied = False
    export_url = None
    
    if request.method == 'POST':
        form = HealthIndicatorReportForm(request.POST)
        if form.is_valid():
            filter_applied = True
            health_indicator = form.cleaned_data['health_indicator']
            min_value = form.cleaned_data['min_value']
            max_value = form.cleaned_data['max_value']
            include_latest_only = form.cleaned_data['include_latest_only']
            
            # Build query
            if include_latest_only:
                # Get latest record for each user, then filter
                user_records = {}
                all_records = HealthRecord.objects.select_related('user__profile').order_by('user', '-recorded_at')
                
                for record in all_records:
                    if record.user_id not in user_records:
                        user_records[record.user_id] = record
                
                # Filter records based on the indicator value
                filtered_records = []
                for record in user_records.values():
                    field_value = getattr(record, health_indicator, None)
                    if field_value is not None and min_value <= float(field_value) <= max_value:
                        filtered_records.append(record)
                        
            else:
                # Get all records that match the criteria
                filter_kwargs = {f'{health_indicator}__range': (min_value, max_value)}
                filtered_records = HealthRecord.objects.select_related('user__profile').filter(
                    **filter_kwargs
                ).order_by('user__username', '-recorded_at')
            
            # Prepare data for display
            filtered_users = []
            seen_users = set() if include_latest_only else None
            
            for record in filtered_records:
                user = record.user
                profile = user.profile
                
                if include_latest_only and user.id in seen_users:
                    continue
                if include_latest_only:
                    seen_users.add(user.id)
                
                # Get the indicator value and status
                indicator_value = getattr(record, health_indicator, None)
                indicator_status = get_indicator_status(record, health_indicator, profile)
                
                user_data = {
                    'user': user,
                    'profile': profile,
                    'record': record,
                    'indicator_value': indicator_value,
                    'indicator_status': indicator_status,
                    'recorded_at': record.recorded_at,
                }
                filtered_users.append(user_data)
            
            # Generate export URL with parameters
            export_params = {
                'health_indicator': health_indicator,
                'min_value': min_value,
                'max_value': max_value,
                'include_latest_only': include_latest_only,
            }
            export_url = f"?export=excel&" + "&".join([f"{k}={v}" for k, v in export_params.items()])
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        return export_health_report_excel(request)
    
    context = {
        'form': form,
        'filtered_users': filtered_users,
        'filter_applied': filter_applied,
        'export_url': export_url,
        'total_users': len(filtered_users),
    }
    
    return render(request, 'health_app/health_report.html', context)


def get_indicator_status(record, indicator, profile):
    """Get status information for a specific health indicator"""
    status_methods = {
        'bmi': record.get_bmi_status,
        'blood_pressure_systolic': record.get_blood_pressure_status,
        'blood_pressure_diastolic': record.get_blood_pressure_status,
        'fat_percent': record.get_fat_percent_status,
        'visceral_fat': record.get_visceral_fat_status,
        'muscle_percent': record.get_muscle_percent_status,
        'waist': record.get_waist_status,
        'cholesterol': record.get_cholesterol_status,
        'ldl': record.get_ldl_status,
        'hdl': record.get_hdl_status,
        'fbs': record.get_fbs_status,
        'triglycerides': record.get_triglycerides_status,
    }
    
    if indicator in status_methods:
        return status_methods[indicator]()
    
    # For indicators without specific status methods
    return {
        'text': 'ไม่มีข้อมูลสถานะ',
        'color': '#6c757d'
    }


def export_health_report_excel(request):
    """Export health report to Excel file"""
    # Check permissions
    if not request.user.is_staff and not request.user.profile.role in ['admin', 'superuser']:
        messages.error(request, 'คุณไม่มีสิทธิ์ในการส่งออกข้อมูล')
        return redirect('health_app:health_report')
    
    # Get filter parameters
    health_indicator = request.GET.get('health_indicator')
    min_value = float(request.GET.get('min_value', 0))
    max_value = float(request.GET.get('max_value', 999999))
    include_latest_only = request.GET.get('include_latest_only') == 'True'
    
    # Apply same filtering logic as the view
    if include_latest_only:
        user_records = {}
        all_records = HealthRecord.objects.select_related('user__profile').order_by('user', '-recorded_at')
        
        for record in all_records:
            if record.user_id not in user_records:
                user_records[record.user_id] = record
        
        filtered_records = []
        for record in user_records.values():
            field_value = getattr(record, health_indicator, None)
            if field_value is not None and min_value <= float(field_value) <= max_value:
                filtered_records.append(record)
    else:
        filter_kwargs = {f'{health_indicator}__range': (min_value, max_value)}
        filtered_records = HealthRecord.objects.select_related('user__profile').filter(
            **filter_kwargs
        ).order_by('user__username', '-recorded_at')
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Health Report'
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Get indicator display name
    indicator_choices = dict(HealthIndicatorReportForm.HEALTH_INDICATOR_CHOICES)
    indicator_name = indicator_choices.get(health_indicator, health_indicator)
    
    # Set headers
    headers = [
        'ลำดับ',
        'ชื่อผู้ใช้',
        'ชื่อ',
        'นามสกุล',
        'เพศ',
        'อายุ',
        f'{indicator_name}',
        'สถานะ',
        'วันที่บันทึก'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Write data
    seen_users = set() if include_latest_only else None
    row = 2
    
    for record in filtered_records:
        user = record.user
        profile = user.profile
        
        if include_latest_only and user.id in seen_users:
            continue
        if include_latest_only:
            seen_users.add(user.id)
        
        indicator_value = getattr(record, health_indicator, None)
        indicator_status = get_indicator_status(record, health_indicator, profile)
        
        worksheet.cell(row=row, column=1, value=row - 1)  # Serial number
        worksheet.cell(row=row, column=2, value=user.username)
        worksheet.cell(row=row, column=3, value=profile.firstname)
        worksheet.cell(row=row, column=4, value=profile.lastname)
        worksheet.cell(row=row, column=5, value='ชาย' if profile.gender == 'male' else 'หญิง')
        worksheet.cell(row=row, column=6, value=profile.age)
        worksheet.cell(row=row, column=7, value=float(indicator_value) if indicator_value else 'N/A')
        worksheet.cell(row=row, column=8, value=indicator_status.get('text', 'ไม่มีข้อมูล') if indicator_status else 'ไม่มีข้อมูล')
        worksheet.cell(row=row, column=9, value=record.recorded_at.strftime('%Y-%m-%d %H:%M'))
        
        row += 1
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        # Skip merged cells that don't have column_letter attribute
        if hasattr(column[0], 'column_letter'):
            column_letter = column[0].column_letter
        else:
            continue
        
        for cell in column:
            try:
                if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'health_report_{indicator_name}_{min_value}-{max_value}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    workbook.save(response)
    
    # Log activity
    log_health_activity(
        request.user, 
        'health_record_export', 
        f'Exported health report for {health_indicator}: {min_value}-{max_value}',
        request
    )
    
    return response


@login_required
def health_status_report_view(request):
    """Health status report with Normal/Abnormal categorization"""
    # Check if user has admin privileges
    if not request.user.is_staff and not request.user.profile.role in ['admin', 'superuser']:
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงรายงานสุขภาพ')
        return redirect('health_app:dashboard')
    
    form = HealthStatusReportForm()
    categorized_users = {}
    filter_applied = False
    export_url = None
    statistics = {}
    
    if request.method == 'POST':
        form = HealthStatusReportForm(request.POST)
        if form.is_valid():
            filter_applied = True
            health_indicator = form.cleaned_data['health_indicator']
            status_filter = form.cleaned_data['status_filter']
            include_latest_only = form.cleaned_data['include_latest_only']
            
            # Get records based on latest_only preference
            if include_latest_only:
                user_records = {}
                all_records = HealthRecord.objects.select_related('user__profile').order_by('user', '-recorded_at')
                
                for record in all_records:
                    if record.user_id not in user_records:
                        user_records[record.user_id] = record
                
                filtered_records = list(user_records.values())
            else:
                filtered_records = HealthRecord.objects.select_related('user__profile').order_by('user__username', '-recorded_at')
            
            # Categorize users by health status
            normal_users = []
            abnormal_users = []
            no_data_users = []
            
            seen_users = set() if include_latest_only else None
            
            for record in filtered_records:
                user = record.user
                profile = user.profile
                
                if include_latest_only and user.id in seen_users:
                    continue
                if include_latest_only:
                    seen_users.add(user.id)
                
                # Get health status for the selected indicator
                health_status = get_health_status_classification(record, health_indicator, profile)
                
                if health_status is None:
                    no_data_users.append({
                        'user': user,
                        'profile': profile,
                        'record': record,
                        'status': 'ไม่มีข้อมูล',
                        'value': 'N/A',
                        'recorded_at': record.recorded_at,
                    })
                    continue
                
                user_data = {
                    'user': user,
                    'profile': profile,
                    'record': record,
                    'status': health_status['status'],
                    'status_detail': health_status,
                    'value': health_status['value'],
                    'recorded_at': record.recorded_at,
                }
                
                if health_status['is_normal']:
                    normal_users.append(user_data)
                else:
                    abnormal_users.append(user_data)
            
            # Apply status filter
            if status_filter == 'normal':
                categorized_users = {'normal': normal_users}
            elif status_filter == 'abnormal':
                categorized_users = {'abnormal': abnormal_users}
            else:  # 'all'
                categorized_users = {
                    'normal': normal_users,
                    'abnormal': abnormal_users,
                    'no_data': no_data_users
                }
            
            # Calculate statistics
            total_users = len(normal_users) + len(abnormal_users) + len(no_data_users)
            statistics = {
                'total': total_users,
                'normal': len(normal_users),
                'abnormal': len(abnormal_users),
                'no_data': len(no_data_users),
                'normal_percentage': round((len(normal_users) / total_users * 100) if total_users > 0 else 0, 1),
                'abnormal_percentage': round((len(abnormal_users) / total_users * 100) if total_users > 0 else 0, 1),
            }
            
            # Generate export URL
            export_params = {
                'health_indicator': health_indicator,
                'status_filter': status_filter,
                'include_latest_only': include_latest_only,
            }
            export_url = f"?export=excel&" + "&".join([f"{k}={v}" for k, v in export_params.items()])
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        return export_health_status_excel(request)
    
    context = {
        'form': form,
        'categorized_users': categorized_users,
        'filter_applied': filter_applied,
        'export_url': export_url,
        'statistics': statistics,
    }
    
    return render(request, 'health_app/health_status_report.html', context)


def get_health_status_classification(record, indicator, profile):
    """Classify health indicator as normal or abnormal"""
    
    # Get the actual value
    if indicator == 'blood_pressure':
        # Special case for blood pressure - check both systolic and diastolic
        systolic = record.blood_pressure_systolic
        diastolic = record.blood_pressure_diastolic
        bp_status = record.get_blood_pressure_status()
        
        # Consider normal if status text is 'เหมาะสม' or 'ปกติ'
        is_normal = bp_status['text'] in ['เหมาะสม', 'ปกติ']
        
        return {
            'value': f"{systolic}/{diastolic}",
            'status': bp_status['text'],
            'color': bp_status['color'],
            'is_normal': is_normal,
            'normal_range': bp_status['normal_range'],
        }
    
    # Get field value
    field_value = getattr(record, indicator, None)
    if field_value is None:
        return None
    
    # Get status based on indicator type
    status_methods = {
        'bmi': record.get_bmi_status,
        'fat_percent': record.get_fat_percent_status,
        'visceral_fat': record.get_visceral_fat_status,
        'muscle_percent': record.get_muscle_percent_status,
        'waist': record.get_waist_status,
        'cholesterol': record.get_cholesterol_status,
        'ldl': record.get_ldl_status,
        'hdl': record.get_hdl_status,
        'fbs': record.get_fbs_status,
        'triglycerides': record.get_triglycerides_status,
    }
    
    if indicator in status_methods:
        status_result = status_methods[indicator]()
        if status_result is None:
            return None
        
        # Determine if status is normal based on status text
        normal_keywords = ['ปกติ', 'เหมาะสม', 'ดีมาก', 'สูง'] # 'สูง' for muscle percent is good
        abnormal_keywords = ['ผิดปกติ', 'เสี่ยง', 'อันตราย', 'ต่ำ', 'สูงกว่าปกติ', 'เริ่มอ้วน', 'อ้วน', 'เบาหวาน', 'สูงมาก']
        
        status_text = status_result['text']
        
        # Special handling for muscle percent where 'สูง' and 'สูงมาก' are good
        if indicator == 'muscle_percent':
            is_normal = status_text in ['ปกติ', 'สูง', 'สูงมาก']
        else:
            is_normal = any(keyword in status_text for keyword in normal_keywords) and \
                       not any(keyword in status_text for keyword in abnormal_keywords)
        
        return {
            'value': float(field_value),
            'status': status_text,
            'color': status_result['color'],
            'is_normal': is_normal,
            'normal_range': status_result.get('normal_range', ''),
        }
    
    # For indicators without specific status methods
    return {
        'value': float(field_value),
        'status': 'ไม่มีข้อมูลสถานะ',
        'color': '#6c757d',
        'is_normal': True,  # Default to normal if no status method
        'normal_range': '',
    }


def export_health_status_excel(request):
    """Export health status report to Excel"""
    # Check permissions
    if not request.user.is_staff and not request.user.profile.role in ['admin', 'superuser']:
        messages.error(request, 'คุณไม่มีสิทธิ์ในการส่งออกข้อมูล')
        return redirect('health_app:health_status_report')
    
    # Get parameters
    health_indicator = request.GET.get('health_indicator')
    status_filter = request.GET.get('status_filter', 'all')
    include_latest_only = request.GET.get('include_latest_only') == 'True'
    
    # Get data using same logic as the view
    if include_latest_only:
        user_records = {}
        all_records = HealthRecord.objects.select_related('user__profile').order_by('user', '-recorded_at')
        
        for record in all_records:
            if record.user_id not in user_records:
                user_records[record.user_id] = record
        
        filtered_records = list(user_records.values())
    else:
        filtered_records = HealthRecord.objects.select_related('user__profile').order_by('user__username', '-recorded_at')
    
    # Categorize data
    normal_users = []
    abnormal_users = []
    no_data_users = []
    
    seen_users = set() if include_latest_only else None
    
    for record in filtered_records:
        user = record.user
        profile = user.profile
        
        if include_latest_only and user.id in seen_users:
            continue
        if include_latest_only:
            seen_users.add(user.id)
        
        health_status = get_health_status_classification(record, health_indicator, profile)
        
        user_data = {
            'user': user,
            'profile': profile,
            'record': record,
            'status': health_status['status'] if health_status else 'ไม่มีข้อมูล',
            'value': health_status['value'] if health_status else 'N/A',
            'is_normal': health_status['is_normal'] if health_status else False,
            'normal_range': health_status['normal_range'] if health_status else '',
        }
        
        if health_status is None:
            no_data_users.append(user_data)
        elif health_status['is_normal']:
            normal_users.append(user_data)
        else:
            abnormal_users.append(user_data)
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    normal_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    abnormal_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
    no_data_fill = PatternFill(start_color='6c757d', end_color='6c757d', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Get indicator display name
    indicator_choices = dict(HealthStatusReportForm.HEALTH_INDICATOR_CHOICES)
    indicator_name = indicator_choices.get(health_indicator, health_indicator)
    
    # Create worksheets based on filter
    if status_filter == 'all':
        # Create separate sheets for each category
        sheets_data = [
            ('Normal', normal_users, normal_fill),
            ('Abnormal', abnormal_users, abnormal_fill),
            ('No Data', no_data_users, no_data_fill)
        ]
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create summary sheet
        summary_sheet = workbook.create_sheet('Summary')
        summary_headers = ['Category', 'Count', 'Percentage']
        
        for col, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = center_alignment
        
        total = len(normal_users) + len(abnormal_users) + len(no_data_users)
        summary_data = [
            ('ปกติ', len(normal_users), f"{(len(normal_users)/total*100):.1f}%" if total > 0 else "0%"),
            ('ผิดปกติ', len(abnormal_users), f"{(len(abnormal_users)/total*100):.1f}%" if total > 0 else "0%"),
            ('ไม่มีข้อมูล', len(no_data_users), f"{(len(no_data_users)/total*100):.1f}%" if total > 0 else "0%"),
            ('รวม', total, '100%')
        ]
        
        for row, (category, count, percentage) in enumerate(summary_data, 2):
            summary_sheet.cell(row=row, column=1, value=category)
            summary_sheet.cell(row=row, column=2, value=count)
            summary_sheet.cell(row=row, column=3, value=percentage)
        
        # Auto-adjust summary sheet columns
        for column in summary_sheet.columns:
            # Skip merged cells that don't have column_letter attribute
            if hasattr(column[0], 'column_letter'):
                max_length = max(len(str(cell.value)) for cell in column if hasattr(cell, 'value'))
                summary_sheet.column_dimensions[column[0].column_letter].width = max_length + 2
    
    else:
        # Single sheet for filtered results
        if status_filter == 'normal':
            sheets_data = [('Normal Results', normal_users, normal_fill)]
        elif status_filter == 'abnormal':
            sheets_data = [('Abnormal Results', abnormal_users, abnormal_fill)]
        
        workbook.remove(workbook.active)
    
    # Create data sheets
    for sheet_name, users_data, fill_color in sheets_data:
        if not users_data:  # Skip empty sheets
            continue
            
        worksheet = workbook.create_sheet(sheet_name)
        
        # Headers
        headers = [
            'ลำดับ', 'ชื่อผู้ใช้', 'ชื่อ', 'นามสกุล', 'เพศ', 'อายุ',
            f'{indicator_name}', 'ค่าปกติ', 'สถานะ', 'วันที่บันทึก'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = fill_color
            cell.alignment = center_alignment
        
        # Data rows
        for row, user_data in enumerate(users_data, 2):
            user = user_data['user']
            profile = user_data['profile']
            record = user_data['record']
            
            worksheet.cell(row=row, column=1, value=row - 1)
            worksheet.cell(row=row, column=2, value=user.username)
            worksheet.cell(row=row, column=3, value=profile.firstname)
            worksheet.cell(row=row, column=4, value=profile.lastname)
            worksheet.cell(row=row, column=5, value='ชาย' if profile.gender == 'male' else 'หญิง')
            worksheet.cell(row=row, column=6, value=profile.age)
            worksheet.cell(row=row, column=7, value=str(user_data['value']))
            worksheet.cell(row=row, column=8, value=user_data['normal_range'])
            worksheet.cell(row=row, column=9, value=user_data['status'])
            worksheet.cell(row=row, column=10, value=record.recorded_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = 0
            # Skip merged cells that don't have column_letter attribute
            if hasattr(column[0], 'column_letter'):
                column_letter = column[0].column_letter
            else:
                continue
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'health_status_report_{indicator_name}_{status_filter}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook
    workbook.save(response)
    
    # Log activity
    log_health_activity(
        request.user, 
        'health_record_export', 
        f'Exported health status report for {health_indicator} ({status_filter})',
        request
    )
    
    return response


@login_required
def admin_health_report_view(request):
    """Comprehensive admin health report with advanced search, graphs, comparison tables, and CRUD"""
    # Check admin privileges
    if not request.user.is_staff and not request.user.profile.role in ['admin', 'superuser']:
        messages.error(request, 'คุณไม่มีสิทธิ์เข้าถึงรายงานผู้ดูแลระบบ')
        return redirect('health_app:dashboard')
    
    form = AdminHealthReportForm()
    search_results = []
    graph_base64 = None
    comparison_data = {}
    latest_overview = []
    user_analysis_data = []
    all_records = []
    search_applied = False
    statistics = {}
    
    if request.method == 'POST':
        form = AdminHealthReportForm(request.POST)
        if form.is_valid():
            search_applied = True
            
            # Get form data
            firstname = form.cleaned_data.get('firstname', '').strip()
            lastname = form.cleaned_data.get('lastname', '').strip()
            gender = form.cleaned_data.get('gender')
            age_min = form.cleaned_data.get('age_min')
            age_max = form.cleaned_data.get('age_max')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            health_index = form.cleaned_data.get('health_index')
            
            # Build user filters
            user_filters = Q()
            
            if firstname:
                user_filters &= Q(profile__firstname__icontains=firstname)
            if lastname:
                user_filters &= Q(profile__lastname__icontains=lastname)
            if gender:
                user_filters &= Q(profile__gender=gender)
            if age_min:
                user_filters &= Q(profile__age__gte=age_min)
            if age_max:
                user_filters &= Q(profile__age__lte=age_max)
            
            # Get users that match criteria
            if user_filters:
                matched_users = User.objects.select_related('profile').filter(user_filters)
            else:
                matched_users = User.objects.select_related('profile').all()
            
            # Build health record filters
            record_filters = Q(user__in=matched_users)
            
            if date_from:
                record_filters &= Q(recorded_at__date__gte=date_from)
            if date_to:
                record_filters &= Q(recorded_at__date__lte=date_to)
            
            # Get health records
            health_records = HealthRecord.objects.select_related('user__profile').filter(record_filters).order_by('-recorded_at')
            
            # Process search results
            search_results = process_search_results(matched_users, health_records, health_index)
            
            # Generate progress graph for filtered results
            if health_records.exists():
                graph_base64 = generate_admin_progress_graph(health_records, health_index)
            
            # Generate comparison data for each user
            comparison_data = generate_admin_comparison_data(matched_users, health_index)
            
            # Generate latest overview for all matched users
            latest_overview = generate_latest_overview(matched_users)
            
            # Generate individual user analysis data (history.html style)
            user_analysis_data = generate_user_analysis_data(matched_users, health_index)
            
            # Get all records for CRUD table (DESC order)
            all_records = list(health_records[:100])  # Limit to 100 for performance
            
            # Calculate statistics
            statistics = calculate_admin_statistics(search_results, health_records)
    
    # Handle AJAX requests for record operations
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'DELETE':
            record_id = request.GET.get('record_id')
            try:
                record = HealthRecord.objects.get(id=record_id)
                record.delete()
                log_health_activity(request.user, 'health_record_delete', f'Deleted record ID: {record_id}', request)
                return JsonResponse({'success': True, 'message': 'ลบข้อมูลสำเร็จ'})
            except HealthRecord.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ไม่พบข้อมูล'})
        
        elif request.method == 'PUT':
            # Handle record updates via AJAX
            record_id = request.GET.get('record_id')
            try:
                record = HealthRecord.objects.get(id=record_id)
                # Update logic would go here
                return JsonResponse({'success': True, 'message': 'อัปเดตข้อมูลสำเร็จ'})
            except HealthRecord.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ไม่พบข้อมูล'})
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        section = request.GET.get('section', 'all')
        form_data = {}  # Initialize form_data
        
        # If GET export request, process search parameters from GET
        if request.method == 'GET' and any(request.GET.get(param) for param in ['firstname', 'lastname', 'gender', 'age_min', 'age_max', 'date_from', 'date_to', 'health_index']):
            # Reprocess search with GET parameters
            form = AdminHealthReportForm(request.GET)
            if form.is_valid():
                # Get form data from GET request
                firstname = form.cleaned_data.get('firstname', '').strip()
                lastname = form.cleaned_data.get('lastname', '').strip()
                gender = form.cleaned_data.get('gender')
                age_min = form.cleaned_data.get('age_min')
                age_max = form.cleaned_data.get('age_max')
                date_from = form.cleaned_data.get('date_from')
                date_to = form.cleaned_data.get('date_to')
                health_index = form.cleaned_data.get('health_index')
                
                # Build user filters
                user_filters = Q()
                if firstname:
                    user_filters &= Q(profile__firstname__icontains=firstname)
                if lastname:
                    user_filters &= Q(profile__lastname__icontains=lastname)
                if gender:
                    user_filters &= Q(profile__gender=gender)
                if age_min:
                    user_filters &= Q(profile__age__gte=age_min)
                if age_max:
                    user_filters &= Q(profile__age__lte=age_max)
                
                # Get users that match criteria
                if user_filters:
                    matched_users = User.objects.select_related('profile').filter(user_filters)
                else:
                    matched_users = User.objects.select_related('profile').all()
                
                # Build health record filters
                record_filters = Q(user__in=matched_users)
                if date_from:
                    record_filters &= Q(recorded_at__date__gte=date_from)
                if date_to:
                    record_filters &= Q(recorded_at__date__lte=date_to)
                
                # Get health records
                health_records = HealthRecord.objects.select_related('user__profile').filter(record_filters).order_by('-recorded_at')
                
                # Regenerate data for export
                search_results = process_search_results(matched_users, health_records, health_index)
                comparison_data = generate_admin_comparison_data(matched_users, health_index)
                latest_overview = generate_latest_overview(matched_users)
                all_records = list(health_records[:100])  # Convert to list to avoid slicing issues
                statistics = calculate_admin_statistics(search_results, health_records)
                
                # Prepare form data for export
                form_data = {
                    'firstname': firstname,
                    'lastname': lastname,
                    'gender': gender,
                    'age_min': age_min,
                    'age_max': age_max,
                    'date_from': date_from,
                    'date_to': date_to,
                    'health_index': health_index,
                }
        
        return export_admin_report_excel(request, search_results, all_records, comparison_data, latest_overview, statistics, section, form_data)
    
    # Handle PDF export
    if request.GET.get('export') == 'pdf':
        section = request.GET.get('section', 'health_overview')
        form_data = {}  # Initialize form_data
        
        # If GET export request, process search parameters from GET
        if request.method == 'GET' and any(request.GET.get(param) for param in ['firstname', 'lastname', 'gender', 'age_min', 'age_max', 'date_from', 'date_to', 'health_index']):
            # Reprocess search with GET parameters (same logic as Excel export)
            form = AdminHealthReportForm(request.GET)
            if form.is_valid():
                # Get form data from GET request
                firstname = form.cleaned_data.get('firstname', '').strip()
                lastname = form.cleaned_data.get('lastname', '').strip()
                gender = form.cleaned_data.get('gender')
                age_min = form.cleaned_data.get('age_min')
                age_max = form.cleaned_data.get('age_max')
                date_from = form.cleaned_data.get('date_from')
                date_to = form.cleaned_data.get('date_to')
                health_index = form.cleaned_data.get('health_index')
                
                # Build user filters
                user_filters = Q()
                if firstname:
                    user_filters &= Q(profile__firstname__icontains=firstname)
                if lastname:
                    user_filters &= Q(profile__lastname__icontains=lastname)
                if gender:
                    user_filters &= Q(profile__gender=gender)
                if age_min:
                    user_filters &= Q(profile__age__gte=age_min)
                if age_max:
                    user_filters &= Q(profile__age__lte=age_max)
                
                # Get users that match criteria
                if user_filters:
                    matched_users = User.objects.select_related('profile').filter(user_filters)
                else:
                    matched_users = User.objects.select_related('profile').all()
                
                # Build health record filters
                record_filters = Q(user__in=matched_users)
                if date_from:
                    record_filters &= Q(recorded_at__date__gte=date_from)
                if date_to:
                    record_filters &= Q(recorded_at__date__lte=date_to)
                
                # Get health records
                health_records = HealthRecord.objects.select_related('user__profile').filter(record_filters).order_by('-recorded_at')
                
                # Process data
                search_results = process_search_results(matched_users, health_records, health_index)
                comparison_data = generate_admin_comparison_data(matched_users, health_index)
                latest_overview = generate_latest_overview(matched_users)
                all_records = list(health_records[:100])
                
                # Statistics
                statistics = calculate_admin_statistics(search_results, health_records)
                
                form_data = {
                    'firstname': firstname,
                    'lastname': lastname,
                    'gender': gender,
                    'age_min': age_min,
                    'age_max': age_max,
                    'date_from': date_from,
                    'date_to': date_to,
                    'health_index': health_index,
                }
        else:
            # Use current search state if available
            if search_applied:
                search_results = search_results
                comparison_data = comparison_data
                latest_overview = latest_overview
                all_records = all_records
                statistics = statistics
            else:
                # Generate default data for PDF export
                all_users = User.objects.select_related('profile').all()
                all_health_records = HealthRecord.objects.select_related('user__profile').all()
                search_results = process_search_results(all_users, all_health_records, None)
                comparison_data = generate_admin_comparison_data(all_users, None)
                latest_overview = generate_latest_overview(all_users)
                all_records = list(all_health_records[:100])
                statistics = calculate_admin_statistics(search_results, all_health_records)
        
        return export_health_overview_pdf(request, latest_overview, section, form_data)
    
    # Generate export URLs with search parameters
    export_url = None
    export_urls = {}
    
    # Always generate export URLs (both for search applied and default state)
    export_params = {}
    if search_applied and form.is_valid():
        # Build export parameters from POST data
        for field_name, field in form.fields.items():
            value = form.cleaned_data.get(field_name)
            if value:
                if hasattr(value, 'strftime'):  # Handle date fields
                    export_params[field_name] = value.strftime('%Y-%m-%d')
                else:
                    export_params[field_name] = value
    
    # Create base export URL
    base_params = "&".join([f"{k}={v}" for k, v in export_params.items()])
    export_url = f"?export=excel&{base_params}" if base_params else "?export=excel"
    
    # Create section-specific export URLs
    export_urls = {
        'complete': export_url,
        'graph': f"?export=excel&section=graph&{base_params}" if base_params else "?export=excel&section=graph",
        'comparison': f"?export=excel&section=comparison&{base_params}" if base_params else "?export=excel&section=comparison",
        'overview': f"?export=excel&section=overview&{base_params}" if base_params else "?export=excel&section=overview",
        'records': f"?export=excel&section=records&{base_params}" if base_params else "?export=excel&section=records",
        'complete_history': f"?export=excel&section=complete_history&{base_params}" if base_params else "?export=excel&section=complete_history",
        'selected_user': f"?export=excel&section=selected_user&{base_params}" if base_params else "?export=excel&section=selected_user",
        'crud_all': f"?export=excel&section=crud_all&{base_params}" if base_params else "?export=excel&section=crud_all",
    }
    
    # Handle export for sections that don't require search data, but if selected_user is requested without search, provide default
    if request.GET.get('export') == 'excel' and section in ['all', 'graph', 'comparison', 'overview', 'records', 'complete_history', 'current_user', 'crud_all']:
        # For non-search dependent sections, call export with basic data
        if not search_applied:
            # Generate default data for export
            all_users = User.objects.select_related('profile').all()
            all_health_records = HealthRecord.objects.select_related('user__profile').all()
            search_results = process_search_results(all_users, all_health_records, None)
            comparison_data = generate_admin_comparison_data(all_users, None)
            latest_overview = generate_latest_overview(all_users)
            all_records = list(all_health_records[:100])
            statistics = calculate_admin_statistics(search_results, all_health_records)
            
            return export_admin_report_excel(request, search_results, all_records, comparison_data, latest_overview, statistics, section, {})
    
    # Ensure we have some default data for export even if no search is applied
    if not search_applied:
        # Get users and records but don't slice QuerySets that will be passed to functions expecting QuerySet methods
        all_users_queryset = User.objects.select_related('profile').all()
        recent_records_queryset = HealthRecord.objects.select_related('user__profile').all().order_by('-recorded_at')
        
        # Limit users to first 50 for processing (convert to list after slicing)
        all_users = list(all_users_queryset[:50])
        
        # For records, keep QuerySet for functions that need .count(), .exists(), etc.
        # Only slice when needed for display (all_records)
        search_results = process_search_results(all_users, recent_records_queryset, 'bmi')
        all_records = list(recent_records_queryset[:100])  # Only convert to list for display
        
        # Pass user list and QuerySet to functions
        comparison_data = generate_admin_comparison_data(all_users, 'bmi')
        latest_overview = generate_latest_overview(all_users)
        statistics = calculate_admin_statistics(search_results, recent_records_queryset)

    context = {
        'form': form,
        'search_results': search_results,
        'graph_base64': graph_base64,
        'comparison_data': comparison_data,
        'latest_overview': latest_overview,
        'user_analysis_data': user_analysis_data if search_applied else [],
        'all_records': all_records,
        'search_applied': search_applied,
        'statistics': statistics,
        'export_url': export_url,
        'export_urls': export_urls,
    }
    
    return render(request, 'health_app/admin_health_report.html', context)


def process_search_results(users, health_records, health_index):
    """Process search results to get user data with health metrics"""
    results = []
    
    for user in users:
        # Handle both QuerySet and list for health_records
        if hasattr(health_records, 'filter'):
            # QuerySet - use filter method
            latest_record = health_records.filter(user=user).first()
            total_records_count = health_records.filter(user=user).count()
        else:
            # List - use list comprehension
            user_records = [r for r in health_records if r.user == user]
            latest_record = user_records[0] if user_records else None
            total_records_count = len(user_records)
        
        if latest_record:
            # Get health index value and status
            health_value = get_health_index_value(latest_record, health_index)
            health_status = get_health_index_status(latest_record, health_index, user.profile)
            
            results.append({
                'user': user,
                'profile': user.profile,
                'latest_record': latest_record,
                'health_value': health_value,
                'health_status': health_status,
                'total_records': total_records_count,
            })
    
    return results


def get_health_index_value(record, health_index):
    """Get the value for a specific health index"""
    if health_index == 'blood_pressure':
        return f"{record.blood_pressure_systolic}/{record.blood_pressure_diastolic}"
    
    # Safety check for None health_index
    if health_index is None or not isinstance(health_index, str):
        return 'N/A'
        
    field_value = getattr(record, health_index, None)
    if field_value is not None:
        return float(field_value)
    return 'N/A'


def get_health_index_status(record, health_index, profile):
    """Get the status for a specific health index"""
    # Safety check for None health_index
    if health_index is None or not isinstance(health_index, str):
        return {'text': 'N/A', 'color': 'secondary'}
        
    status_methods = {
        'bmi': record.get_bmi_status,
        'blood_pressure': record.get_blood_pressure_status,
        'fat_percent': record.get_fat_percent_status,
        'visceral_fat': record.get_visceral_fat_status,
        'muscle_percent': record.get_muscle_percent_status,
        'waist': record.get_waist_status,
        'cholesterol': record.get_cholesterol_status,
        'ldl': record.get_ldl_status,
        'hdl': record.get_hdl_status,
        'fbs': record.get_fbs_status,
        'triglycerides': record.get_triglycerides_status,
    }
    
    if health_index in status_methods:
        return status_methods[health_index]()
    
    return {
        'text': 'ไม่มีข้อมูลสถานะ',
        'color': '#6c757d'
    }


def generate_admin_progress_graph(health_records, health_index):
    """Generate progress graph for admin report"""
    if not health_records.exists():
        return None
    
    plt.figure(figsize=(14, 8))
    plt.rcParams['font.family'] = 'DejaVu Sans'
    
    # Group records by user and plot separate lines
    users_data = {}
    for record in health_records.order_by('recorded_at'):
        user_id = record.user_id
        if user_id not in users_data:
            users_data[user_id] = {
                'dates': [],
                'values': [],
                'username': record.user.username
            }
        
        # Get value for the health index
        if health_index == 'blood_pressure':
            value = record.blood_pressure_systolic  # Use systolic for graph
        else:
            value = getattr(record, health_index, None)
        
        if value is not None:
            users_data[user_id]['dates'].append(record.recorded_at)
            users_data[user_id]['values'].append(float(value))
    
    # Plot data for each user (limit to top 10 users for readability)
    colors = plt.cm.Set3(np.linspace(0, 1, min(len(users_data), 10)))
    
    for i, (user_id, data) in enumerate(list(users_data.items())[:10]):
        if data['dates'] and data['values']:
            plt.plot(data['dates'], data['values'], 
                    marker='o', linestyle='-', linewidth=2, markersize=6, 
                    label=data['username'], color=colors[i], alpha=0.8)
    
    # Get health index display name
    index_names = {
        'bmi': 'BMI',
        'blood_pressure': 'Blood Pressure (Systolic)',
        'fat_percent': 'Fat Percent (%)',
        'visceral_fat': 'Visceral Fat',
        'muscle_percent': 'Muscle Percent (%)',
        'waist': 'Waist Circumference (cm)',
        'cholesterol': 'Cholesterol (mg/dL)',
        'ldl': 'LDL (mg/dL)',
        'hdl': 'HDL (mg/dL)',
        'fbs': 'FBS (mg/dL)',
        'triglycerides': 'Triglycerides (mg/dL)',
    }
    
    index_name = index_names.get(health_index, health_index)
    
    plt.title(f'Health Progress Report - {index_name}', fontsize=16, fontweight='bold')
    plt.xlabel('Date', fontsize=12)
    plt.ylabel(index_name, fontsize=12)
    plt.xticks(rotation=45)
    plt.grid(True, alpha=0.3)
    
    if len(users_data) <= 10:
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    
    plt.tight_layout()
    
    # Convert to base64
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    
    return image_base64


def generate_admin_comparison_data(users, health_index):
    """Generate comparison data (first vs last vs normal) for each user"""
    comparison_data = []
    
    for user in users:
        records = HealthRecord.objects.filter(user=user).order_by('recorded_at')
        
        if records.count() >= 2:
            first_record = records.first()
            last_record = records.last()
            
            # Get values
            first_value = get_health_index_value(first_record, health_index)
            last_value = get_health_index_value(last_record, health_index)
            
            # Get normal range
            normal_range = get_normal_range_for_index(health_index, user.profile)
            
            # Calculate change
            change = calculate_change(first_value, last_value)
            
            # Get current status
            current_status = get_health_index_status(last_record, health_index, user.profile)
            
            comparison_data.append({
                'user': user,
                'profile': user.profile,
                'first_value': first_value,
                'last_value': last_value,
                'normal_range': normal_range,
                'change': change,
                'status': current_status,
                'first_date': first_record.recorded_at,
                'last_date': last_record.recorded_at,
            })
    
    return comparison_data


def get_normal_range_for_index(health_index, profile):
    """Get normal range text for health index"""
    ranges = {
        'bmi': '18.5 - 22.9',
        'blood_pressure': '<120/<80',
        'fat_percent': '20-29.9%' if profile.gender == 'female' else '10-19.9%',
        'visceral_fat': '1 - 9',
        'waist': f'≈{float(profile.age or 30) * 2.5:.1f} cm',  # Approximate with fallback
        'cholesterol': '<200 mg/dL',
        'ldl': '<130 mg/dL',
        'hdl': '>50 mg/dL' if profile.gender == 'female' else '>40 mg/dL',
        'fbs': '<100 mg/dL',
        'triglycerides': '<150 mg/dL',
    }
    
    # For muscle percent, depends on age and gender
    if health_index == 'muscle_percent':
        age = profile.age
        if profile.gender == 'female':
            if 18 <= age <= 39:
                return '24.3 - 30.3%'
            elif 40 <= age <= 59:
                return '24.1 - 30.1%'
            else:
                return '23.9 - 29.9%'
        else:
            if 18 <= age <= 39:
                return '33.3 - 39.3%'
            elif 40 <= age <= 59:
                return '33.1 - 39.1%'
            else:
                return '32.9 - 38.9%'
    
    return ranges.get(health_index, 'N/A')


def calculate_change(first_value, last_value):
    """Calculate change between first and last values"""
    if isinstance(first_value, str) and '/' in first_value:
        # Blood pressure case
        return f"{first_value} → {last_value}"
    
    try:
        first_num = float(first_value)
        last_num = float(last_value)
        diff = last_num - first_num
        
        if diff > 0:
            return f"+{diff:.1f}"
        elif diff < 0:
            return f"{diff:.1f}"
        else:
            return "ไม่เปลี่ยนแปลง"
    except (ValueError, TypeError):
        return "ไม่สามารถคำนวณได้"


def generate_latest_overview(users):
    """Generate comprehensive overview of latest values for all users with detailed health analysis"""
    overview = []
    
    for user in users:
        latest_record = HealthRecord.objects.filter(user=user).first()
        
        if latest_record:
            profile = user.profile
            
            # Get all health statuses (comprehensive)
            bmi_status = latest_record.get_bmi_status()
            bp_status = latest_record.get_blood_pressure_status()
            fat_status = latest_record.get_fat_percent_status()
            visceral_status = latest_record.get_visceral_fat_status()
            muscle_status = latest_record.get_muscle_percent_status()
            waist_status = latest_record.get_waist_status()
            
            # Additional health indicators (if available)
            cholesterol_status = latest_record.get_cholesterol_status()
            ldl_status = latest_record.get_ldl_status()
            hdl_status = latest_record.get_hdl_status()
            fbs_status = latest_record.get_fbs_status()
            triglycerides_status = latest_record.get_triglycerides_status()
            
            # Generate comprehensive health overview like history view
            health_overview_analysis = generate_health_overview(
                latest_record, profile, bmi_status, fat_status, visceral_status, muscle_status
            )
            
            # Count normal vs abnormal indicators (extended)
            normal_count = 0
            total_count = 0
            
            all_statuses = [
                bmi_status, bp_status, fat_status, visceral_status, 
                muscle_status, waist_status, cholesterol_status, 
                ldl_status, hdl_status, fbs_status, triglycerides_status
            ]
            
            status_details = {}
            
            for i, status in enumerate(all_statuses):
                if status:
                    total_count += 1
                    status_names = [
                        'bmi', 'blood_pressure', 'fat_percent', 'visceral_fat',
                        'muscle_percent', 'waist', 'cholesterol', 'ldl', 'hdl', 'fbs', 'triglycerides'
                    ]
                    status_key = status_names[i]
                    status_details[status_key] = status
                    
                    # Enhanced normal detection
                    if status['text'] in ['ปกติ', 'เหมาะสม', 'ดีมาก', 'สูง'] and \
                       status_key in ['muscle_percent'] or \
                       status['text'] in ['ปกติ', 'เหมาะสม', 'ดีมาก'] and \
                       status_key not in ['muscle_percent']:
                        normal_count += 1
            
            health_score = (normal_count / total_count * 100) if total_count > 0 else 0
            
            # Generate detailed health metrics summary
            health_metrics = {
                'bmi': {
                    'value': latest_record.bmi,
                    'status': bmi_status,
                    'unit': ''
                },
                'blood_pressure': {
                    'value': f"{latest_record.blood_pressure_systolic}/{latest_record.blood_pressure_diastolic}",
                    'status': bp_status,
                    'unit': 'mmHg'
                },
                'fat_percent': {
                    'value': latest_record.fat_percent,
                    'status': fat_status,
                    'unit': '%'
                },
                'visceral_fat': {
                    'value': latest_record.visceral_fat,
                    'status': visceral_status,
                    'unit': ''
                },
                'muscle_percent': {
                    'value': latest_record.muscle_percent,
                    'status': muscle_status,
                    'unit': '%'
                },
                'waist': {
                    'value': latest_record.waist,
                    'status': waist_status,
                    'unit': 'cm'
                }
            }
            
            # Add optional metrics if available
            if latest_record.cholesterol:
                health_metrics['cholesterol'] = {
                    'value': latest_record.cholesterol,
                    'status': cholesterol_status,
                    'unit': 'mg/dL'
                }
            
            if latest_record.ldl:
                health_metrics['ldl'] = {
                    'value': latest_record.ldl,
                    'status': ldl_status,
                    'unit': 'mg/dL'
                }
            
            if latest_record.hdl:
                health_metrics['hdl'] = {
                    'value': latest_record.hdl,
                    'status': hdl_status,
                    'unit': 'mg/dL'
                }
            
            if latest_record.fbs:
                health_metrics['fbs'] = {
                    'value': latest_record.fbs,
                    'status': fbs_status,
                    'unit': 'mg/dL'
                }
            
            if latest_record.triglycerides:
                health_metrics['triglycerides'] = {
                    'value': latest_record.triglycerides,
                    'status': triglycerides_status,
                    'unit': 'mg/dL'
                }
            
            # Risk assessment
            risk_level = 'low'
            if health_score < 50:
                risk_level = 'high'
            elif health_score < 75:
                risk_level = 'medium'
            
            # Generate comparison if user has multiple records
            comparison_analysis = None
            user_records = HealthRecord.objects.filter(user=user).order_by('recorded_at')
            if user_records.count() >= 2:
                first_record = user_records.first()
                comparison_analysis = generate_comparison_table(first_record, latest_record, profile)
            
            overview.append({
                'user': user,
                'profile': user.profile,
                'latest_record': latest_record,
                'health_score': health_score,
                'normal_count': normal_count,
                'total_count': total_count,
                'risk_level': risk_level,
                
                # Status details for individual metrics
                'bmi_status': bmi_status,
                'bp_status': bp_status,
                'fat_status': fat_status,
                'visceral_status': visceral_status,
                'muscle_status': muscle_status,
                'waist_status': waist_status,
                'cholesterol_status': cholesterol_status,
                'ldl_status': ldl_status,
                'hdl_status': hdl_status,
                'fbs_status': fbs_status,
                'triglycerides_status': triglycerides_status,
                
                # Comprehensive health data
                'health_metrics': health_metrics,
                'health_overview_analysis': health_overview_analysis,
                'comparison_analysis': comparison_analysis,
                'status_details': status_details,
                
                # Demographic info
                'age': profile.age,
                'gender': profile.gender,
                'gender_display': 'ชาย' if profile.gender == 'male' else 'หญิง',
                
                # Record info
                'recorded_date': latest_record.recorded_at,
                'total_records': user_records.count(),
            })
    
    # Sort by health score descending, then by risk level
    overview.sort(key=lambda x: (x['health_score'], x['risk_level'] == 'low'), reverse=True)
    
    return overview


def generate_user_analysis_data(users, health_index):
    """Generate comprehensive individual user analysis data similar to history.html"""
    user_analysis_data = []
    
    for user in users:
        records = HealthRecord.objects.filter(user=user).order_by('recorded_at')
        
        if records.exists():
            # Calculate user's age for comparison analysis
            age = user.profile.age or 30
            
            # Generate individual graph for this user
            graph_base64 = generate_individual_user_graph(records, health_index, user)
            
            # Generate comparison data for this user
            comparison_data = generate_individual_user_comparison(records, user.profile)
            
            # Generate summary info
            summary_info = generate_individual_user_summary(records.last(), user.profile)
            
            # Calculate health score for this user
            latest_record = records.last()
            health_score = calculate_individual_health_score(latest_record)
            
            user_analysis_data.append({
                'user': user,
                'profile': user.profile,
                'age': age,
                'records': records,
                'records_count': records.count(),
                'graph_base64': graph_base64,
                'comparison_data': comparison_data,
                'summary_info': summary_info,
                'health_score': health_score,
            })
    
    return user_analysis_data


def generate_individual_user_graph(records, health_index, user):
    """Generate individual progress graph for a specific user"""
    if not records.exists():
        return None
    
    plt.figure(figsize=(10, 6))
    plt.rcParams['font.family'] = 'DejaVu Sans'
    
    dates = []
    values = []
    
    for record in records:
        if health_index == 'blood_pressure':
            value = record.blood_pressure_systolic
        else:
            value = getattr(record, health_index, None)
        
        if value is not None:
            dates.append(record.recorded_at)
            values.append(float(value))
    
    if dates and values:
        plt.plot(dates, values, marker='o', linestyle='-', linewidth=2, markersize=8, 
                color='#4a90e2', alpha=0.8)
        
        # Add trend line
        if len(values) > 1:
            z = np.polyfit(range(len(values)), values, 1)
            p = np.poly1d(z)
            plt.plot(dates, p(range(len(values))), "r--", alpha=0.5, linewidth=1)
    
    # Get health index display name
    index_names = {
        'bmi': 'BMI',
        'blood_pressure': 'Blood Pressure (Systolic)',
        'fat_percent': 'Fat Percent (%)',
        'visceral_fat': 'Visceral Fat',
        'muscle_percent': 'Muscle Percent (%)',
        'waist': 'Waist Circumference (cm)',
        'cholesterol': 'Cholesterol (mg/dL)',
        'ldl': 'LDL (mg/dL)',
        'hdl': 'HDL (mg/dL)',
        'fbs': 'FBS (mg/dL)',
        'triglycerides': 'Triglycerides (mg/dL)',
    }
    
    index_name = index_names.get(health_index, health_index)
    
    plt.title(f'{index_name} Progress - {user.profile.firstname} {user.profile.lastname}', 
              fontsize=14, fontweight='bold')
    plt.xlabel('Date', fontsize=10)
    plt.ylabel(index_name, fontsize=10)
    plt.xticks(rotation=45, fontsize=9)
    plt.yticks(fontsize=9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    # Convert to base64
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=80, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    
    return image_base64


def generate_individual_user_comparison(records, profile):
    """Generate comparison data for individual user (first vs last values)"""
    if records.count() < 2:
        return None
    
    first_record = records.first()
    last_record = records.last()
    
    comparison_data = {}
    
    # Define metrics to compare
    metrics = {
        'bmi': {'unit': '', 'method': 'get_bmi_status'},
        'fat_percent': {'unit': '%', 'method': 'get_fat_percent_status'},
        'visceral_fat': {'unit': '', 'method': 'get_visceral_fat_status'},
        'muscle_percent': {'unit': '%', 'method': 'get_muscle_percent_status'},
        'blood_pressure_systolic': {'unit': '', 'method': 'get_blood_pressure_status'},
        'blood_pressure_diastolic': {'unit': '', 'method': 'get_blood_pressure_status'},
        'waist': {'unit': ' cm', 'method': 'get_waist_status'},
        'cholesterol': {'unit': ' mg/dL', 'method': 'get_cholesterol_status'},
        'ldl': {'unit': ' mg/dL', 'method': 'get_ldl_status'},
        'hdl': {'unit': ' mg/dL', 'method': 'get_hdl_status'},
        'fbs': {'unit': ' mg/dL', 'method': 'get_fbs_status'},
        'triglycerides': {'unit': ' mg/dL', 'method': 'get_triglycerides_status'},
    }
    
    for metric, config in metrics.items():
        first_value = getattr(first_record, metric, None)
        last_value = getattr(last_record, metric, None)
        
        if first_value is not None and last_value is not None:
            # Get status using the method from the last record
            status_method = getattr(last_record, config['method'], None)
            status = status_method() if status_method else {'text': 'ไม่มีข้อมูล', 'color': '#6c757d'}
            
            # Calculate change
            try:
                change = float(last_value) - float(first_value)
                if change > 0:
                    change_text = f"+{change:.1f}"
                elif change < 0:
                    change_text = f"{change:.1f}"
                else:
                    change_text = "ไม่เปลี่ยนแปลง"
            except (ValueError, TypeError):
                change_text = "ไม่สามารถคำนวณได้"
            
            comparison_data[metric] = {
                'start_value': f"{first_value:.1f}" if isinstance(first_value, (int, float)) else first_value,
                'end_value': f"{last_value:.1f}" if isinstance(last_value, (int, float)) else last_value,
                'change': change_text,
                'text': status['text'],
                'color': status['color'],
            }
    
    return comparison_data


def generate_individual_user_summary(latest_record, profile):
    """Generate health summary for individual user"""
    summary_info = {}
    
    # Get all health analysis for the latest record
    bmi_status = latest_record.get_bmi_status()
    fat_status = latest_record.get_fat_percent_status()
    visceral_status = latest_record.get_visceral_fat_status()
    muscle_status = latest_record.get_muscle_percent_status()
    overview = generate_health_overview(latest_record, profile, bmi_status, fat_status, visceral_status, muscle_status)
    
    # Extract key summaries
    for item in overview:
        metric_name = item['text'].lower()
        if 'bmi' in metric_name or 'ดัชนี' in metric_name:
            summary_info['bmi'] = item['text']
        elif 'เอว' in metric_name or 'waist' in metric_name:
            summary_info['waist'] = item['text']
        elif 'ความดัน' in metric_name or 'blood pressure' in metric_name:
            summary_info['blood_pressure_systolic'] = item['text']
            summary_info['blood_pressure_diastolic'] = item['text']
        elif 'ไขมัน' in metric_name and 'เปอร์เซ็นต์' in metric_name:
            summary_info['fat_percent'] = item['text']
        elif 'ไขมันในช่องท้อง' in metric_name or 'visceral' in metric_name:
            summary_info['visceral_fat'] = item['text']
        elif 'กล้ามเนื้อ' in metric_name or 'muscle' in metric_name:
            summary_info['muscle_percent'] = item['text']
        elif 'คอเลสเตอรอล' in metric_name or 'cholesterol' in metric_name:
            summary_info['cholesterol'] = item['text']
        elif 'ldl' in metric_name.lower():
            summary_info['ldl'] = item['text']
        elif 'hdl' in metric_name.lower():
            summary_info['hdl'] = item['text']
        elif 'น้ำตาล' in metric_name or 'fbs' in metric_name.lower():
            summary_info['fbs'] = item['text']
        elif 'ไตรกลี' in metric_name or 'triglycerides' in metric_name:
            summary_info['triglycerides'] = item['text']
    
    return summary_info


def calculate_individual_health_score(latest_record):
    """Calculate health score for individual user"""
    if not latest_record:
        return 0
    
    # Get all health statuses
    bmi_status = latest_record.get_bmi_status()
    bp_status = latest_record.get_blood_pressure_status()
    fat_status = latest_record.get_fat_percent_status()
    visceral_status = latest_record.get_visceral_fat_status()
    muscle_status = latest_record.get_muscle_percent_status()
    waist_status = latest_record.get_waist_status()
    
    # Count normal vs abnormal indicators
    normal_count = 0
    total_count = 0
    
    statuses = [bmi_status, bp_status, fat_status, visceral_status, muscle_status, waist_status]
    
    for status in statuses:
        if status:
            total_count += 1
            if status['text'] in ['ปกติ', 'เหมาะสม', 'ดีมาก']:
                normal_count += 1
    
    return (normal_count / total_count * 100) if total_count > 0 else 0


def calculate_admin_statistics(search_results, health_records):
    """Calculate statistics for admin report"""
    # Handle empty or None inputs
    if not search_results:
        search_results = []
    if not health_records:
        health_records = []
        
    total_users = len(search_results)
    # Handle both QuerySet and list for health_records
    try:
        total_records = health_records.count() if hasattr(health_records, 'count') else len(health_records)
    except Exception:
        total_records = 0
    
    # Calculate average records per user
    avg_records_per_user = total_records / total_users if total_users > 0 else 0
    
    # Count users by health status
    healthy_users = 0
    at_risk_users = 0
    
    for result in search_results:
        if result['health_status']:
            if result['health_status']['text'] in ['ปกติ', 'เหมาะสม', 'ดีมาก']:
                healthy_users += 1
            else:
                at_risk_users += 1
    
    # Calculate date range
    date_range = "N/A"
    if hasattr(health_records, 'exists'):
        # QuerySet - use QuerySet methods
        if health_records.exists():
            latest_date = health_records.order_by('-recorded_at').first().recorded_at
            earliest_date = health_records.order_by('recorded_at').first().recorded_at
            date_range = f"{earliest_date.strftime('%Y-%m-%d')} to {latest_date.strftime('%Y-%m-%d')}"
    else:
        # List - use list methods
        if health_records:
            # Sort by recorded_at to get earliest and latest
            sorted_records = sorted(health_records, key=lambda x: x.recorded_at)
            earliest_date = sorted_records[0].recorded_at
            latest_date = sorted_records[-1].recorded_at
            date_range = f"{earliest_date.strftime('%Y-%m-%d')} to {latest_date.strftime('%Y-%m-%d')}"
    
    return {
        'total_users': total_users,
        'total_records': total_records,
        'avg_records_per_user': round(avg_records_per_user, 1),
        'healthy_users': healthy_users,
        'at_risk_users': at_risk_users,
        'healthy_percentage': round(healthy_users / total_users * 100, 1) if total_users > 0 else 0,
        'date_range': date_range,
    }


def export_admin_report_excel(request, search_results, all_records, comparison_data=None, latest_overview=None, statistics=None, section='all', form_data=None):
    """Export comprehensive admin report to Excel with multiple sheets including graphs and analysis"""
    print(f"DEBUG: Excel export called with section={section}, search_results count={len(search_results) if search_results else 0}, form_data={form_data}")
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Border, Side
    
    workbook = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Colors for status indicators
    success_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    warning_fill = PatternFill(start_color='ffc107', end_color='ffc107', fill_type='solid')
    danger_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Determine which sections to include
    include_sections = {
        'summary': section in ['all'],
        'graph': section in ['all', 'graph'],
        'comparison': section in ['all', 'comparison'],
        'overview': section in ['all', 'overview'],
        'records': section in ['all', 'records'],
        'complete_history': section in ['complete_history'],
        'selected_user': section in ['selected_user'],
        'crud_all': section in ['crud_all']
    }
    
    # Get form data for filtering context from both POST and GET parameters
    form_data = {}
    if request.method == 'POST':
        form = AdminHealthReportForm(request.POST)
        if form.is_valid():
            for field_name, field in form.fields.items():
                form_data[field_name] = form.cleaned_data.get(field_name)
    else:
        # Handle GET parameters (from export URLs)
        form = AdminHealthReportForm(request.GET)
        if form.is_valid():
            for field_name, field in form.fields.items():
                value = form.cleaned_data.get(field_name)
                if value:
                    form_data[field_name] = value
        else:
            # Fallback: extract parameters directly from GET if form validation fails
            for key in ['firstname', 'lastname', 'gender', 'age_min', 'age_max', 'date_from', 'date_to', 'health_index']:
                value = request.GET.get(key)
                if value:
                    form_data[key] = value
    
    # Special handler for CRUD-only export to avoid creating unnecessary sheets
    if section == 'crud_all':
        workbook = openpyxl.Workbook()
        
        # Remove default sheet and create only CRUD sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create CRUD sheet
        crud_sheet = workbook.create_sheet('All CRUD Records')
        
        # Define styles for CRUD sheet
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        center_alignment = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Title and description
        crud_sheet['A1'] = 'All Health Records - CRUD Operations Export'
        crud_sheet['A1'].font = Font(bold=True, size=16)
        crud_sheet.merge_cells('A1:P1')
        
        crud_sheet['A2'] = f'Export Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        crud_sheet['A2'].font = Font(italic=True)
        
        # Get all health records
        all_crud_records = HealthRecord.objects.select_related('user__profile').all().order_by('-recorded_at')
        
        crud_sheet['A3'] = f'Total Records: {all_crud_records.count()}'
        crud_sheet['A3'].font = Font(bold=True, color='0066CC')
        
        # Headers matching the CRUD table exactly with all health metrics
        crud_headers = [
            'ID', 'User', 'Name', 'BMI', 'Fat%', 'Visceral', 'Muscle%', 
            'BP', 'Waist', 'Weight (kg)', 'Height (cm)', 'Cholesterol', 
            'LDL', 'HDL', 'FBS', 'Recorded Date'
        ]
        
        # Write headers
        for col, header in enumerate(crud_headers, 1):
            cell = crud_sheet.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Populate CRUD data
        crud_row = 6
        for record in all_crud_records:
            profile = record.user.profile
            
            # Format BP as "systolic/diastolic"
            bp_value = f"{record.blood_pressure_systolic or ''}/{record.blood_pressure_diastolic or ''}"
            if bp_value == "/":
                bp_value = ""
            
            data = [
                record.id,
                record.user.username,
                f"{profile.firstname or ''} {profile.lastname or ''}".strip(),
                float(record.bmi) if record.bmi else '',
                float(record.fat_percent) if record.fat_percent else '',
                float(record.visceral_fat) if record.visceral_fat else '',
                float(record.muscle_percent) if record.muscle_percent else '',
                bp_value,
                float(record.waist) if record.waist else '',
                float(record.weight) if record.weight else '',
                float(record.height) if record.height else '',
                float(record.cholesterol) if record.cholesterol else '',
                float(record.ldl) if record.ldl else '',
                float(record.hdl) if record.hdl else '',
                float(record.fbs) if record.fbs else '',
                record.recorded_at.strftime('%Y-%m-%d %H:%M') if record.recorded_at else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = crud_sheet.cell(row=crud_row, column=col, value=value)
                cell.border = border
                
                # Simple alternating row colors
                if crud_row % 2 == 0:
                    cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            crud_row += 1
        
        # Auto-adjust column widths
        for column in crud_sheet.columns:
            max_length = 0
            if hasattr(column[0], 'column_letter'):
                column_letter = column[0].column_letter
            else:
                continue
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            crud_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Return the CRUD-only workbook
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'admin_health_all_crud_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response
    
    # Get health records and other data for export
    health_index = form_data.get('health_index', 'bmi')
    
    # Use passed data or regenerate if not provided
    if comparison_data is None or latest_overview is None:
        matched_users = [result['user'] for result in search_results]
        comparison_data = generate_admin_comparison_data(matched_users, health_index)
        latest_overview = generate_latest_overview(matched_users)
    
    if statistics is None:
        statistics = calculate_admin_statistics(search_results, all_records)
    
    # Sheet 1: Executive Summary (only for complete exports)
    if include_sections['summary']:
        exec_sheet = workbook.create_sheet('Executive Summary')
        
        # Add title and metadata
        exec_sheet.merge_cells('A1:F1')
        title_cell = exec_sheet['A1']
        title_cell.value = 'Health Progress Admin Report - Executive Summary'
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_alignment
        
        # Add report metadata
        exec_sheet['A3'] = 'Report Generated:'
        exec_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        exec_sheet['A4'] = 'Health Index Analyzed:'
        exec_sheet['B4'] = health_index.upper()
        exec_sheet['A5'] = 'Search Criteria:'
        criteria_text = ', '.join([f"{k}: {v}" for k, v in form_data.items() if v])
        exec_sheet['B5'] = criteria_text if criteria_text else 'No filters applied'
        
        # Add statistics summary
        exec_sheet['A7'] = 'SUMMARY STATISTICS'
        exec_sheet['A7'].font = Font(bold=True)
        
        # Handle cases where statistics might be None or incomplete
        if statistics and isinstance(statistics, dict):
            stats_data = [
                ['Total Users Found:', statistics.get('total_users', 0)],
                ['Total Health Records:', statistics.get('total_records', 0)],
                ['Average Records per User:', statistics.get('avg_records_per_user', 0)],
                ['Healthy Users:', statistics.get('healthy_users', 0)],
                ['At-Risk Users:', statistics.get('at_risk_users', 0)],
                ['Health Percentage:', f"{statistics.get('healthy_percentage', 0)}%"],
                ['Data Date Range:', statistics.get('date_range', 'N/A')]
            ]
        else:
            # Default statistics when none provided
            stats_data = [
                ['Total Users Found:', 0],
                ['Total Health Records:', 0],
                ['Average Records per User:', 0],
                ['Healthy Users:', 0],
                ['At-Risk Users:', 0],
                ['Health Percentage:', '0%'],
                ['Data Date Range:', 'N/A']
            ]
        
        for i, (label, value) in enumerate(stats_data, 8):
            exec_sheet[f'A{i}'] = label
            exec_sheet[f'B{i}'] = value
            exec_sheet[f'A{i}'].font = Font(bold=True)
    
    # Sheet 2: Health Progress Graph Data
    if include_sections['graph']:
        graph_sheet = workbook.create_sheet('Health Progress Graph Data')
        
        # Add graph data export
        graph_sheet['A1'] = f'Health Progress Data - {health_index.upper()}'
        graph_sheet['A1'].font = Font(bold=True, size=14)
        
        # Generate and embed graph image
        try:
            # Create the graph and save it temporarily
            health_records = HealthRecord.objects.filter(user__in=[r['user'] for r in search_results]).order_by('-recorded_at')
            graph_base64 = generate_admin_progress_graph(health_records, health_index)
            
            if graph_base64:
                # Decode base64 and save as temporary image
                import base64
                import tempfile
                import os
                
                graph_data = base64.b64decode(graph_base64)
                
                # Create temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
                    temp_file.write(graph_data)
                    temp_file_path = temp_file.name
                
                # Add image to Excel
                try:
                    img = Image(temp_file_path)
                    img.width = 800  # Adjust size as needed
                    img.height = 400
                    graph_sheet.add_image(img, 'A3')
                    
                    # Clean up temporary file
                    os.unlink(temp_file_path)
                except Exception as e:
                    # If image insertion fails, add note
                    graph_sheet['A3'] = f'Graph image could not be embedded: {str(e)}'
                    try:
                        os.unlink(temp_file_path)
                    except:
                        pass
            else:
                graph_sheet['A3'] = 'No graph data available for the selected criteria'
        except Exception as e:
            graph_sheet['A3'] = f'Graph generation failed: {str(e)}'
        
        # Create graph data table (starting at row 25 to leave space for graph image)
        graph_sheet['A25'] = 'Detailed Graph Data:'
        graph_sheet['A25'].font = Font(bold=True)
        
        graph_headers = ['Username', 'Full Name', 'Date', f'{health_index.upper()} Value', 'Status']
        for col, header in enumerate(graph_headers, 1):
            cell = graph_sheet.cell(row=27, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Populate graph data
        graph_row = 28
        for record in all_records:
            user = record.user
            profile = user.profile
            health_value = get_health_index_value(record, health_index)
            health_status = get_health_index_status(record, health_index, profile)
            
            graph_sheet.cell(row=graph_row, column=1, value=user.username).border = border
            graph_sheet.cell(row=graph_row, column=2, value=f"{profile.firstname} {profile.lastname}").border = border
            graph_sheet.cell(row=graph_row, column=3, value=record.recorded_at.strftime('%Y-%m-%d')).border = border
            
            value_cell = graph_sheet.cell(row=graph_row, column=4, value=str(health_value))
            value_cell.border = border
            
            status_cell = graph_sheet.cell(row=graph_row, column=5, value=health_status['text'] if health_status else 'N/A')
            status_cell.border = border
            
            # Color code status
            if health_status and health_status['text'] in ['ปกติ', 'เหมาะสม', 'ดีมาก']:
                status_cell.fill = success_fill
                status_cell.font = Font(color='FFFFFF')
            elif health_status and health_status['text'] in ['สูงกว่าปกติ', 'เริ่มอ้วน']:
                status_cell.fill = warning_fill
            elif health_status:
                status_cell.fill = danger_fill
                status_cell.font = Font(color='FFFFFF')
            
            graph_row += 1
    
    # Sheet 3: Comparison Table (First vs Last vs Normal)
    if include_sections['comparison']:
        comparison_sheet = workbook.create_sheet('Comparison Analysis')
    
        comparison_sheet['A1'] = f'Health Comparison Analysis - {health_index.upper()}'
        comparison_sheet['A1'].font = Font(bold=True, size=14)
    
        comp_headers = [
            'Username', 'Full Name', 'Gender', 'Age',
            'First Date', 'First Value', 'Last Date', 'Last Value', 
            'Normal Range', 'Change', 'Current Status'
        ]
    
        for col, header in enumerate(comp_headers, 1):
            cell = comparison_sheet.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
    
        # Populate comparison data
        comp_row = 4
        for item in comparison_data:
            user = item['user']
            profile = item['profile']
            
            comparison_sheet.cell(row=comp_row, column=1, value=user.username).border = border
            comparison_sheet.cell(row=comp_row, column=2, value=f"{profile.firstname} {profile.lastname}").border = border
            comparison_sheet.cell(row=comp_row, column=3, value='ชาย' if profile.gender == 'male' else 'หญิง').border = border
            comparison_sheet.cell(row=comp_row, column=4, value=profile.age).border = border
            comparison_sheet.cell(row=comp_row, column=5, value=item['first_date'].strftime('%Y-%m-%d')).border = border
            comparison_sheet.cell(row=comp_row, column=6, value=str(item['first_value'])).border = border
            comparison_sheet.cell(row=comp_row, column=7, value=item['last_date'].strftime('%Y-%m-%d')).border = border
            comparison_sheet.cell(row=comp_row, column=8, value=str(item['last_value'])).border = border
            comparison_sheet.cell(row=comp_row, column=9, value=item['normal_range']).border = border
            comparison_sheet.cell(row=comp_row, column=10, value=item['change']).border = border
            
            status_cell = comparison_sheet.cell(row=comp_row, column=11, value=item['status']['text'])
            status_cell.border = border
            
            # Color code status
            if item['status']['text'] in ['ปกติ', 'เหมาะสม', 'ดีมาก']:
                status_cell.fill = success_fill
                status_cell.font = Font(color='FFFFFF')
            elif item['status']['text'] in ['สูงกว่าปกติ', 'เริ่มอ้วน']:
                status_cell.fill = warning_fill
            else:
                status_cell.fill = danger_fill
                status_cell.font = Font(color='FFFFFF')
            
            comp_row += 1
    
    # Sheet 4: Latest Health Overview
    if include_sections['overview']:
        overview_sheet = workbook.create_sheet('Latest Health Overview')
        
        overview_sheet['A1'] = 'Latest Health Status Overview'
        overview_sheet['A1'].font = Font(bold=True, size=14)
    
        overview_headers = [
            'Username', 'Full Name', 'Gender', 'Age', 'Health Score (%)',
            'Normal Indicators', 'Total Indicators', 'BMI Status', 'BP Status', 
            'Fat % Status', 'Visceral Fat Status', 'Muscle % Status', 'Waist Status',
            'Latest Record Date'
        ]
        
        for col, header in enumerate(overview_headers, 1):
            cell = overview_sheet.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Populate overview data
        overview_row = 4
        for item in latest_overview:
            user = item['user']
            profile = item['profile']
            record = item['latest_record']
            
            overview_sheet.cell(row=overview_row, column=1, value=user.username).border = border
            overview_sheet.cell(row=overview_row, column=2, value=f"{profile.firstname} {profile.lastname}").border = border
            overview_sheet.cell(row=overview_row, column=3, value='ชาย' if profile.gender == 'male' else 'หญิง').border = border
            overview_sheet.cell(row=overview_row, column=4, value=profile.age).border = border
        
        # Health score with color coding
        score_cell = overview_sheet.cell(row=overview_row, column=5, value=round(item['health_score'], 1))
        score_cell.border = border
        if item['health_score'] >= 80:
            score_cell.fill = success_fill
            score_cell.font = Font(color='FFFFFF')
        elif item['health_score'] >= 60:
            score_cell.fill = warning_fill
        else:
            score_cell.fill = danger_fill
            score_cell.font = Font(color='FFFFFF')
        
        overview_sheet.cell(row=overview_row, column=6, value=item['normal_count']).border = border
        overview_sheet.cell(row=overview_row, column=7, value=item['total_count']).border = border
        
        # Status columns
        statuses = [
            item.get('bmi_status', {}),
            item.get('bp_status', {}),
            item.get('fat_status', {}),
            record.get_visceral_fat_status() if record else {},
            record.get_muscle_percent_status() if record else {},
            record.get_waist_status() if record else {}
        ]
        
        for i, status in enumerate(statuses, 8):
            status_cell = overview_sheet.cell(row=overview_row, column=i, 
                                            value=status.get('text', 'N/A') if status else 'N/A')
            status_cell.border = border
            
            if status and status.get('text') in ['ปกติ', 'เหมาะสม', 'ดีมาก']:
                status_cell.fill = success_fill
                status_cell.font = Font(color='FFFFFF')
            elif status and status.get('text') in ['สูงกว่าปกติ', 'เริ่มอ้วน']:
                status_cell.fill = warning_fill
        
        overview_sheet.cell(row=overview_row, column=14, value=record.recorded_at.strftime('%Y-%m-%d %H:%M')).border = border
        
        overview_row += 1
    
    # Sheet 5: All Records (CRUD Data) - Enhanced
    if include_sections['records']:
        records_sheet = workbook.create_sheet('All Health Records')
    
    records_sheet['A1'] = 'Complete Health Records Database'
    records_sheet['A1'].font = Font(bold=True, size=14)
    
    record_headers = [
        'Record ID', 'Username', 'First Name', 'Last Name', 'Gender', 'Age',
        'BMI', 'Fat %', 'Visceral Fat', 'Muscle %', 
        'BP Systolic', 'BP Diastolic', 'Waist (cm)', 'Height (cm)', 'Weight (kg)',
        'Cholesterol', 'LDL', 'HDL', 'FBS', 'Triglycerides',
        'BMR', 'Body Age', 'Recorded Date', 'Created Date'
    ]
    
    for col, header in enumerate(record_headers, 1):
        cell = records_sheet.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Populate complete records data
    record_row = 4
    for record in all_records:
        user = record.user
        profile = user.profile
        
        data = [
            record.id, user.username, profile.firstname, profile.lastname,
            'ชาย' if profile.gender == 'male' else 'หญิง', profile.age,
            float(record.bmi), float(record.fat_percent), float(record.visceral_fat), 
            float(record.muscle_percent), record.blood_pressure_systolic, 
            record.blood_pressure_diastolic, float(record.waist), float(record.height), 
            float(record.weight), float(record.cholesterol or 0), float(record.ldl or 0),
            float(record.hdl or 0), float(record.fbs or 0), float(record.triglycerides or 0),
            record.bmr, record.body_age, record.recorded_at.strftime('%Y-%m-%d %H:%M'),
            record.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col, value in enumerate(data, 1):
            cell = records_sheet.cell(row=record_row, column=col, value=value)
            cell.border = border
            if col <= 6:  # User info columns
                cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
        
        record_row += 1
    
    # Sheet 6: Complete Records History (All Users, No Limits)
    if include_sections['complete_history']:
        history_sheet = workbook.create_sheet('Complete Records History')
        
        # Title and description
        history_sheet['A1'] = 'Complete Health Records History - All Users'
        history_sheet['A1'].font = Font(bold=True, size=16)
        history_sheet.merge_cells('A1:X1')
        
        history_sheet['A2'] = f'Export Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        history_sheet['A2'].font = Font(italic=True)
        
        # Get ALL health records without any limits
        all_history_records = HealthRecord.objects.select_related('user__profile').all().order_by('-recorded_at')
        
        # Apply search filters if any
        if form_data:
            history_filters = Q()
            
            # User filters
            if form_data.get('firstname'):
                history_filters &= Q(user__profile__firstname__icontains=form_data['firstname'])
            if form_data.get('lastname'):
                history_filters &= Q(user__profile__lastname__icontains=form_data['lastname'])
            if form_data.get('gender'):
                history_filters &= Q(user__profile__gender=form_data['gender'])
            if form_data.get('age_min'):
                history_filters &= Q(user__profile__age__gte=form_data['age_min'])
            if form_data.get('age_max'):
                history_filters &= Q(user__profile__age__lte=form_data['age_max'])
            
            # Date filters
            if form_data.get('date_from'):
                if isinstance(form_data['date_from'], str):
                    date_from = datetime.strptime(form_data['date_from'], '%Y-%m-%d').date()
                else:
                    date_from = form_data['date_from']
                history_filters &= Q(recorded_at__date__gte=date_from)
            if form_data.get('date_to'):
                if isinstance(form_data['date_to'], str):
                    date_to = datetime.strptime(form_data['date_to'], '%Y-%m-%d').date()
                else:
                    date_to = form_data['date_to']
                history_filters &= Q(recorded_at__date__lte=date_to)
            
            if history_filters:
                all_history_records = all_history_records.filter(history_filters)
        
        # Headers with additional information
        history_headers = [
            'ID', 'Username', 'Email', 'First Name', 'Last Name', 'Gender', 'Age', 'Birth Date',
            'BMI', 'BMI Status', 'Fat %', 'Fat Status', 'Visceral Fat', 'Visceral Status',
            'Muscle %', 'Muscle Status', 'BP Systolic', 'BP Diastolic', 'BP Status',
            'Waist (cm)', 'Waist Status', 'Height (cm)', 'Weight (kg)',
            'Cholesterol', 'LDL', 'HDL', 'FBS', 'Triglycerides',
            'BMR', 'Body Age', 'Health Score', 'Recorded Date', 'Created Date'
        ]
        
        # Write headers
        for col, header in enumerate(history_headers, 1):
            cell = history_sheet.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Summary info
        records_count = all_history_records.count() if hasattr(all_history_records, 'count') else len(all_history_records)
        history_sheet['A3'] = f'Total Records: {records_count}'
        history_sheet['A3'].font = Font(bold=True, color='0066CC')
        
        # Populate all history data
        history_row = 5
        for record in all_history_records:
            user = record.user
            profile = user.profile
            
            # Calculate health statuses
            bmi_status = get_health_index_status(record, 'bmi', profile)
            fat_status = get_health_index_status(record, 'fat_percent', profile)
            visceral_status = get_health_index_status(record, 'visceral_fat', profile)
            muscle_status = get_health_index_status(record, 'muscle_percent', profile)
            bp_status = get_health_index_status(record, 'blood_pressure_systolic', profile)
            waist_status = get_health_index_status(record, 'waist', profile)
            
            # Calculate overall health score
            health_score = calculate_individual_health_score(record)
            
            data = [
                record.id, user.username, user.email or '', profile.firstname or '', 
                profile.lastname or '', 'ชาย' if profile.gender == 'male' else 'หญิง', 
                profile.age or '', profile.birthdate.strftime('%Y-%m-%d') if profile.birthdate else '',
                float(record.bmi), bmi_status, float(record.fat_percent), fat_status,
                float(record.visceral_fat), visceral_status, float(record.muscle_percent), muscle_status,
                record.blood_pressure_systolic or '', record.blood_pressure_diastolic or '', bp_status,
                float(record.waist), waist_status, float(record.height), float(record.weight),
                float(record.cholesterol or 0), float(record.ldl or 0), float(record.hdl or 0),
                float(record.fbs or 0), float(record.triglycerides or 0),
                record.bmr or '', record.body_age or '', health_score,
                record.recorded_at.strftime('%Y-%m-%d %H:%M:%S'),
                record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                cell = history_sheet.cell(row=history_row, column=col, value=value)
                cell.border = border
                
                # Color coding based on health status
                if col in [10, 12, 14, 16, 19, 21]:  # Status columns
                    if 'ปกติ' in str(value) or 'Normal' in str(value) or 'Excellent' in str(value):
                        cell.fill = success_fill
                        cell.font = Font(color='FFFFFF', bold=True)
                    elif 'เสี่ยง' in str(value) or 'Risk' in str(value) or 'High' in str(value):
                        cell.fill = warning_fill
                        cell.font = Font(color='000000', bold=True)
                    elif 'อันตราย' in str(value) or 'Danger' in str(value) or 'Critical' in str(value):
                        cell.fill = danger_fill
                        cell.font = Font(color='FFFFFF', bold=True)
                elif col <= 8:  # User info columns
                    cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            history_row += 1
    
    # Sheet 7: Selected User Complete Records (Based on Search Criteria)
    print(f"DEBUG: include_sections={include_sections}")
    print(f"DEBUG: section={section}")
    if include_sections['selected_user']:
        print("DEBUG: Creating Selected User sheet")
        user_sheet = workbook.create_sheet('Selected User Records')
        
        # Title and description
        user_sheet['A1'] = 'Selected User Complete Health Records'
        user_sheet['A1'].font = Font(bold=True, size=16)
        user_sheet.merge_cells('A1:X1')
        
        user_sheet['A2'] = f'Export Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        user_sheet['A2'].font = Font(italic=True)
        
        # Get selected user records based on search criteria
        selected_user_records = []
        print(f"DEBUG: Selected user export - form_data: {form_data}")
        if form_data and any(form_data.values()):
            # Find the first user that matches the search criteria
            user_filters = Q()
            if form_data.get('firstname'):
                user_filters &= Q(profile__firstname__icontains=form_data['firstname'])
            if form_data.get('lastname'):
                user_filters &= Q(profile__lastname__icontains=form_data['lastname'])
            if form_data.get('gender'):
                user_filters &= Q(profile__gender=form_data['gender'])
            if form_data.get('age_min'):
                user_filters &= Q(profile__age__gte=form_data['age_min'])
            if form_data.get('age_max'):
                user_filters &= Q(profile__age__lte=form_data['age_max'])
            
            print(f"DEBUG: user_filters: {user_filters}")
            if user_filters:
                # Get the first matching user
                selected_user = User.objects.select_related('profile').filter(user_filters).first()
                print(f"DEBUG: Found user: {selected_user}")
                if selected_user:
                    # Get ALL records for this specific user
                    record_filters = Q(user=selected_user)
                    if form_data.get('date_from'):
                        if isinstance(form_data['date_from'], str):
                            date_from = datetime.strptime(form_data['date_from'], '%Y-%m-%d').date()
                        else:
                            date_from = form_data['date_from']
                        record_filters &= Q(recorded_at__date__gte=date_from)
                    if form_data.get('date_to'):
                        if isinstance(form_data['date_to'], str):
                            date_to = datetime.strptime(form_data['date_to'], '%Y-%m-%d').date()
                        else:
                            date_to = form_data['date_to']
                        record_filters &= Q(recorded_at__date__lte=date_to)
                    
                    selected_user_records = HealthRecord.objects.select_related('user__profile').filter(record_filters).order_by('-recorded_at')
                    print(f"DEBUG: Found {selected_user_records.count()} records for user {selected_user.username}")
                    
                    # Add user info to sheet
                    user_sheet['A3'] = f'User: {selected_user.profile.firstname} {selected_user.profile.lastname} ({selected_user.username})'
                    user_sheet['A3'].font = Font(bold=True, color='0066CC')
        
        # If no form data provided, show message and get some sample data
        if not form_data or not any(form_data.values()):
            user_sheet['A3'] = 'No search criteria provided - showing first user with records'
            user_sheet['A3'].font = Font(bold=True, color='FF8C00')
            # Get first user that has health records
            first_user_with_records = User.objects.select_related('profile').filter(
                healthrecord__isnull=False
            ).distinct().first()
            if first_user_with_records:
                selected_user_records = HealthRecord.objects.select_related('user__profile').filter(
                    user=first_user_with_records
                ).order_by('-recorded_at')[:50]  # Limit to 50 records
                user_sheet['A3'] = f'Sample Data - User: {first_user_with_records.profile.firstname} {first_user_with_records.profile.lastname} (First 50 records)'
                user_sheet['A3'].font = Font(bold=True, color='FF8C00')
        
        # If still no specific user found, show message
        elif not selected_user_records or (hasattr(selected_user_records, 'count') and selected_user_records.count() == 0):
            user_sheet['A3'] = 'No specific user found matching search criteria'
            user_sheet['A3'].font = Font(bold=True, color='CC0000')
            selected_user_records = []
        
        # Headers for selected user records - matching the requested format
        user_headers = [
            'First Name', 'Last Name', 'Gender', 'Age', 'BMI', 'Fat %', 'Visceral Fat', 'Muscle %', 
            'BP Systolic', 'BP Diastolic', 'Waist (cm)', 'Height (cm)', 'Weight (kg)', 
            'Cholesterol', 'LDL', 'HDL', 'FBS', 'Triglycerides', 'BMR', 'Body Age', 
            'Recorded Date', 'Created Date'
        ]
        
        # Write headers
        for col, header in enumerate(user_headers, 1):
            cell = user_sheet.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Summary info
        user_sheet['A4'] = f'Total Records: {selected_user_records.count() if hasattr(selected_user_records, "count") else len(selected_user_records)}'
        user_sheet['A4'].font = Font(bold=True, color='0066CC')
        
        # Populate selected user data
        user_row = 6
        record_count = selected_user_records.count() if hasattr(selected_user_records, 'count') else len(selected_user_records)
        print(f"DEBUG: Processing {record_count} records for Excel export")
        for record in selected_user_records:
            profile = record.user.profile
            
            # Calculate age at time of record
            age = profile.age if hasattr(profile, 'age') else ''
            
            data = [
                profile.firstname or '',
                profile.lastname or '',
                profile.gender or '',
                age,
                float(record.bmi) if record.bmi else '',
                float(record.fat_percent) if record.fat_percent else '',
                float(record.visceral_fat) if record.visceral_fat else '',
                float(record.muscle_percent) if record.muscle_percent else '',
                record.blood_pressure_systolic or '',
                record.blood_pressure_diastolic or '',
                float(record.waist) if record.waist else '',
                float(record.height) if record.height else '',
                float(record.weight) if record.weight else '',
                float(record.cholesterol) if record.cholesterol else '',
                float(record.ldl) if record.ldl else '',
                float(record.hdl) if record.hdl else '',
                float(record.fbs) if record.fbs else '',
                float(record.triglycerides) if record.triglycerides else '',
                record.bmr or '',
                record.body_age or '',
                record.recorded_at.strftime('%Y-%m-%d %H:%M:%S'),
                record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                cell = user_sheet.cell(row=user_row, column=col, value=value)
                cell.border = border
                
                # Simple alternating row colors for better readability
                if user_row % 2 == 0:
                    cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            user_row += 1
    
    # Sheet 8: All CRUD Records Export (Complete Table Export)
    if include_sections['crud_all']:
        print("DEBUG: Creating All CRUD Records sheet")
        crud_sheet = workbook.create_sheet('All CRUD Records')
        
        # Title and description
        crud_sheet['A1'] = 'All Health Records - CRUD Operations Export'
        crud_sheet['A1'].font = Font(bold=True, size=16)
        crud_sheet.merge_cells('A1:K1')
        
        crud_sheet['A2'] = f'Export Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        crud_sheet['A2'].font = Font(italic=True)
        
        # Get all health records (not limited to 100 like the display table)
        all_crud_records = HealthRecord.objects.select_related('user__profile').all().order_by('-recorded_at')
        
        crud_sheet['A3'] = f'Total Records: {all_crud_records.count()}'
        crud_sheet['A3'].font = Font(bold=True, color='0066CC')
        
        # Headers matching the CRUD table exactly
        crud_headers = [
            'ID', 'User', 'Name', 'BMI', 'Fat%', 'Visceral', 'Muscle%', 
            'BP', 'Waist', 'Recorded Date'
        ]
        
        # Write headers
        for col, header in enumerate(crud_headers, 1):
            cell = crud_sheet.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Populate CRUD data
        crud_row = 6
        record_count = all_crud_records.count()
        print(f"DEBUG: Processing {record_count} CRUD records for Excel export")
        
        for record in all_crud_records:
            profile = record.user.profile
            
            # Format BP as "systolic/diastolic"
            bp_value = f"{record.blood_pressure_systolic or ''}/{record.blood_pressure_diastolic or ''}"
            if bp_value == "/":
                bp_value = ""
            
            data = [
                record.id,
                record.user.username,
                f"{profile.firstname or ''} {profile.lastname or ''}".strip(),
                float(record.bmi) if record.bmi else '',
                float(record.fat_percent) if record.fat_percent else '',
                float(record.visceral_fat) if record.visceral_fat else '',
                float(record.muscle_percent) if record.muscle_percent else '',
                bp_value,
                float(record.waist) if record.waist else '',
                record.recorded_at.strftime('%Y-%m-%d %H:%M') if record.recorded_at else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = crud_sheet.cell(row=crud_row, column=col, value=value)
                cell.border = border
                
                # Simple alternating row colors for better readability
                if crud_row % 2 == 0:
                    cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            crud_row += 1
    
    # Auto-adjust column widths for all sheets
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            # Skip merged cells that don't have column_letter attribute
            if hasattr(column[0], 'column_letter'):
                column_letter = column[0].column_letter
            else:
                continue
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename based on section
    section_names = {
        'all': 'complete_report',
        'graph': 'progress_graph',
        'comparison': 'comparison_analysis', 
        'overview': 'health_overview',
        'records': 'all_records',
        'complete_history': 'complete_history_all_users',
        'selected_user': 'selected_user_complete_records',
        'crud_all': 'all_crud_records'
    }
    section_name = section_names.get(section, 'report')
    filename = f'admin_health_{section_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    
    # Log activity
    log_health_activity(
        request.user, 
        'admin_report_export', 
        f'Exported admin health report with {len(search_results)} users',
        request
    )
    
    return response


@login_required
def delete_record_view(request, record_id):
    """Delete health record (admin only)"""
    # Check admin privileges
    if not request.user.is_staff and not request.user.profile.role in ['admin', 'superuser']:
        messages.error(request, 'คุณไม่มีสิทธิ์ในการลบข้อมูล')
        return redirect('health_app:admin_health_report')
    
    try:
        record = get_object_or_404(HealthRecord, id=record_id)
        username = record.user.username
        recorded_date = record.recorded_at.strftime('%Y-%m-%d')
        
        record.delete()
        log_health_activity(request.user, 'health_record_delete', f'Deleted record for {username} on {recorded_date}', request)
        messages.success(request, f'ลบข้อมูลสุขภาพของ {username} วันที่ {recorded_date} สำเร็จ')
        
    except HealthRecord.DoesNotExist:
        messages.error(request, 'ไม่พบข้อมูลที่ต้องการลบ')
    
    return redirect('health_app:admin_health_report')


def export_health_overview_pdf(request, latest_overview, section='health_overview', form_data=None):
    """Export Latest Health Overview to PDF with comprehensive analysis"""
    
    # Register Thai fonts
    thai_font = register_thai_fonts()
    
    # Generate filename with user names if available
    filename = "health_overview_report"
    if latest_overview and len(latest_overview) > 0:
        # If single user, include their name
        if len(latest_overview) == 1:
            item = latest_overview[0]
            profile = item.get('profile')
            if profile:
                firstname = getattr(profile, 'firstname', '') or ''
                lastname = getattr(profile, 'lastname', '') or ''
                if firstname or lastname:
                    # Clean names for filename (remove special characters)
                    clean_firstname = ''.join(c for c in firstname if c.isalnum() or c in ' -_').strip()
                    clean_lastname = ''.join(c for c in lastname if c.isalnum() or c in ' -_').strip()
                    user_name = f"{clean_firstname}_{clean_lastname}".replace(' ', '_')
                    if user_name and user_name != '_':
                        filename = f"health_report_{user_name}"
        # If multiple users, include count
        elif len(latest_overview) > 1:
            filename = f"health_report_{len(latest_overview)}_users"
    
    # Add timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_filename = f"{filename}_{timestamp}.pdf"
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{final_filename}"'
    
    # Create the PDF document
    doc = SimpleDocTemplate(response, pagesize=A4, 
                          rightMargin=72, leftMargin=72, 
                          topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles with Thai font support
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        spaceAfter=30,
        textColor=colors.darkblue,
        alignment=TA_CENTER,
        fontName=thai_font
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=20,
        spaceAfter=12,
        textColor=colors.darkblue,
        alignment=TA_LEFT,
        fontName=thai_font
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=18,
        spaceAfter=8,
        textColor=colors.darkred,
        alignment=TA_LEFT,
        fontName=thai_font
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=16,
        spaceAfter=6,
        alignment=TA_LEFT,
        fontName=thai_font
    )
    
    # Title
    title_text = safe_thai_text("Latest Health Overview - Complete Analysis")
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 12))
    
    # Report metadata
    current_time = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    elements.append(Paragraph(safe_thai_text(f"<b>Generated:</b> {current_time}"), normal_style))
    elements.append(Paragraph(safe_thai_text(f"<b>Total Users Analyzed:</b> {len(latest_overview) if latest_overview else 0}"), normal_style))
    
    # Add search criteria if provided
    if form_data:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Search Criteria Applied:", subheading_style))
        for key, value in form_data.items():
            if value:
                display_key = key.replace('_', ' ').title()
                elements.append(Paragraph(f"• <b>{display_key}:</b> {value}", normal_style))
    
    elements.append(Spacer(1, 20))
    
    if not latest_overview:
        elements.append(Paragraph("No health data available for the selected criteria.", normal_style))
        doc.build(elements)
        return response
    
    # Process each user's health overview
    for i, item in enumerate(latest_overview):
        if i > 0:
            elements.append(PageBreak())
        
        # User header - Safe attribute access
        profile = item.get('profile')
        user = item.get('user')
        
        if not profile or not user:
            continue  # Skip if essential data is missing
            
        firstname = getattr(profile, 'firstname', 'Unknown') or 'Unknown'
        lastname = getattr(profile, 'lastname', '') or ''
        username = getattr(user, 'username', 'Unknown') or 'Unknown'
        
        user_name = f"{firstname} {lastname}".strip()
        
        # User title with health score
        health_score = item.get('health_score', 0)
        if health_score >= 80:
            score_color_hex = '#28a745'  # green
        elif health_score >= 60:
            score_color_hex = '#ff7707'  # orange
        else:
            score_color_hex = '#dc3545'  # red
        
        elements.append(Paragraph(safe_thai_text(f"<b>{user_name}</b> ({username})"), heading_style))
        elements.append(Paragraph(safe_thai_text(f"Health Score: <b><font color='{score_color_hex}'>{health_score:.0f}%</font></b>"), normal_style))
        
        # User basic info
        age = item.get('age', 'N/A')
        gender = item.get('gender_display', 'N/A')
        total_records = item.get('total_records', 0)
        recorded_date = item.get('recorded_date')
        
        elements.append(Paragraph(f"<b>Age:</b> {age} years | <b>Gender:</b> {gender} | <b>Total Records:</b> {total_records}", normal_style))
        if recorded_date and hasattr(recorded_date, 'strftime'):
            try:
                date_str = recorded_date.strftime('%B %d, %Y at %I:%M %p')
                elements.append(Paragraph(f"<b>Last Recorded:</b> {date_str}", normal_style))
            except (AttributeError, ValueError):
                elements.append(Paragraph(f"<b>Last Recorded:</b> {str(recorded_date)}", normal_style))
        
        elements.append(Spacer(1, 12))
        
        # Health Metrics Summary
        elements.append(Paragraph(safe_thai_text("Current Health Metrics"), subheading_style))
        
        health_metrics = item.get('health_metrics', {})
        if health_metrics:
            # Create table data for health metrics
            metric_data = []
            metric_data.append([safe_thai_text('Metric'), safe_thai_text('Value'), safe_thai_text('Status'), safe_thai_text('Normal Range')])
            
            for metric_name, metric_info in health_metrics.items():
                # Format metric name
                formatted_name = {
                    'bmi': 'BMI',
                    'blood_pressure': 'Blood Pressure',
                    'fat_percent': 'Body Fat %',
                    'visceral_fat': 'Visceral Fat',
                    'muscle_percent': 'Muscle %',
                    'waist': 'Waist Circumference',
                    'cholesterol': 'Cholesterol',
                    'ldl': 'LDL',
                    'hdl': 'HDL',
                    'fbs': 'Blood Sugar (FBS)',
                    'triglycerides': 'Triglycerides'
                }.get(metric_name, metric_name.title())
                
                # Safe value and status access
                metric_value = metric_info.get('value', 'N/A')
                metric_unit = metric_info.get('unit', '')
                if metric_value is None:
                    metric_value = 'N/A'
                if metric_unit is None:
                    metric_unit = ''
                    
                value = f"{metric_value}{metric_unit}"
                
                # Safe status access
                status_dict = metric_info.get('status', {})
                if status_dict and isinstance(status_dict, dict):
                    status = status_dict.get('text', 'Unknown')
                else:
                    status = 'Unknown'
                
                # Normal ranges
                normal_ranges = {
                    'bmi': '18.5-24.9',
                    'blood_pressure': '<120/80',
                    'fat_percent': 'M: 10-20%, F: 16-25%',
                    'visceral_fat': '1-9',
                    'muscle_percent': 'M: >38%, F: >28%',
                    'waist': 'M: <90cm, F: <80cm',
                    'cholesterol': '<200 mg/dL',
                    'ldl': '<100 mg/dL',
                    'hdl': 'M: >40, F: >50 mg/dL',
                    'fbs': '70-100 mg/dL',
                    'triglycerides': '<150 mg/dL'
                }
                normal_range = normal_ranges.get(metric_name, 'Varies')
                
                metric_data.append([safe_thai_text(formatted_name), safe_thai_text(str(value)), safe_thai_text(str(status)), safe_thai_text(str(normal_range))])
            
            # Create and style the metrics table
            metrics_table = Table(metric_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            # Use safe font names for table headers
            header_font = 'Helvetica-Bold' if thai_font in ['Tahoma', 'Arial', 'THSarabun'] else thai_font + '-Bold' if thai_font != 'Helvetica' else 'Helvetica-Bold'
            content_font = thai_font
            
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), header_font),
                ('FONTNAME', (0, 1), (-1, -1), content_font),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(metrics_table)
            elements.append(Spacer(1, 12))
        
        # Health Analysis Overview
        health_analysis = item.get('health_overview_analysis', [])
        if health_analysis:
            elements.append(Paragraph("Health Analysis", subheading_style))
            
            for j, analysis in enumerate(health_analysis):
                analysis_color = analysis.get('color', 'info')
                status_text = {
                    'success': '✓ Good Status',
                    'warning': '⚠ Caution Needed', 
                    'danger': '✗ High Risk',
                }.get(analysis_color, 'ℹ Information')
                
                # Safe color handling
                if analysis_color == 'success':
                    text_color_hex = '#28a745'  # green
                elif analysis_color == 'warning':
                    text_color_hex = "#ff7707"  # orange
                elif analysis_color == 'danger':
                    text_color_hex = '#dc3545'  # red
                else:
                    text_color_hex = '#6c757d'  # gray
                
                elements.append(Paragraph(f"<b><font color='{text_color_hex}'>{status_text}</font></b>", normal_style))
                elements.append(Paragraph(f"{analysis.get('text', 'No analysis available')}", normal_style))
                elements.append(Spacer(1, 6))
        
        # Overall Health Progress
        normal_count = item.get('normal_count', 0)
        total_count = item.get('total_count', 0)
        
        elements.append(Paragraph("Overall Health Summary", subheading_style))
        elements.append(Paragraph(f"<b>Health Indicators:</b> {normal_count}/{total_count} within normal range", normal_style))
        elements.append(Paragraph(f"<b>Overall Health Score:</b> {health_score:.0f}%", normal_style))
        
        # Risk Level
        risk_level = item.get('risk_level', 'unknown')
        risk_text = {
            'low': 'Low Risk',
            'medium': 'Medium Risk', 
            'high': 'High Risk'
        }.get(risk_level, 'Unknown Risk')
        
        # Safe risk color handling
        if risk_level == 'low':
            risk_color_hex = '#28a745'  # green
        elif risk_level == 'medium':
            risk_color_hex = '#ff7707'  # orange
        elif risk_level == 'high':
            risk_color_hex = '#dc3545'  # red
        else:
            risk_color_hex = '#6c757d'  # gray
            
        elements.append(Paragraph(f"<b>Risk Level:</b> <font color='{risk_color_hex}'>{risk_text}</font>", normal_style))
        
        # Comparison Analysis (if available)
        comparison_analysis = item.get('comparison_analysis')
        if comparison_analysis:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("Progress Comparison (First vs Latest)", subheading_style))
            
            # Create comparison table
            comparison_data = []
            comparison_data.append(['Metric', 'First Record', 'Latest Record', 'Change'])
            
            metrics_to_show = ['bmi', 'fat_percent', 'muscle_percent', 'visceral_fat', 'blood_pressure', 'waist', 'height', 'weight']
            for metric in metrics_to_show:
                if metric in comparison_analysis and comparison_analysis[metric]:
                    metric_info = comparison_analysis[metric]
                    if isinstance(metric_info, dict):
                        formatted_name = {
                            'bmi': 'BMI',
                            'fat_percent': 'Body Fat %',
                            'muscle_percent': 'Muscle %',
                            'visceral_fat': 'Visceral Fat',
                            'blood_pressure': 'Blood Pressure',
                            'waist': 'Waist',
                            'height': 'Height',
                            'weight': 'Weight'
                        }.get(metric, metric.title())
                        
                        first_val = metric_info.get('start_value', 'N/A')
                        latest_val = metric_info.get('end_value', 'N/A')
                        change_text = metric_info.get('text', 'No change')
                        
                        # Convert None to 'N/A'
                        if first_val is None:
                            first_val = 'N/A'
                        if latest_val is None:
                            latest_val = 'N/A'
                        if change_text is None:
                            change_text = 'No change'
                        
                        comparison_data.append([formatted_name, str(first_val), str(latest_val), str(change_text)])
            
            if len(comparison_data) > 1:  # Has data beyond header
                comparison_table = Table(comparison_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2*inch])
                # Use safe font names for comparison table
                comp_header_font = 'Helvetica-Bold' if thai_font in ['Tahoma', 'Arial', 'THSarabun'] else thai_font + '-Bold' if thai_font != 'Helvetica' else 'Helvetica-Bold'
                comp_content_font = thai_font
                
                comparison_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), comp_header_font),
                    ('FONTNAME', (0, 1), (-1, -1), comp_content_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(comparison_table)
        
        elements.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(elements)
    return response
